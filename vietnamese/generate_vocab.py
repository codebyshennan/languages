#!/usr/bin/env python3
"""
generate_vocab.py — Generate ~1,038 B2-level Vietnamese words via Claude API.

Combines with the existing 1,962 words to produce viet_vocab_COMPLETE_3000words.xlsx.

Usage:
    ANTHROPIC_API_KEY=<key> python3 generate_vocab.py

Resumable: saves progress to generate_progress.json after each batch.
Re-running after interruption continues from where it left off.
"""

import json
import os
import re
import time
from collections import Counter
from pathlib import Path

import anthropic
import openpyxl

BASE          = Path(__file__).parent
SOURCE_FILE   = BASE / "viet_vocab_COMPLETE_1962words.xlsx"
OUTPUT_FILE   = BASE / "viet_vocab_COMPLETE_3000words.xlsx"
PROGRESS_FILE = BASE / "generate_progress.json"
SHEET_NAME    = "📚 All Words (Combined)"
TARGET_TOTAL  = 3000
BATCH_SIZE    = 100
MAX_RETRIES   = 3


def normalise(s: str) -> str:
    """Lowercase, strip, collapse internal whitespace. Preserves Vietnamese diacritics."""
    return re.sub(r'\s+', ' ', str(s or "").lower().strip())


def load_existing():
    """Return (rows_list, dedup_set, sorted_cats_list) from the source Excel."""
    wb = openpyxl.load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows, dedup, cats = [], set(), set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        num, part, viet, english, hanzi, cantonese, cat, notes = (list(row) + [""] * 8)[:8]
        rows.append({
            "num":       int(num),
            "part":      str(part      or ""),
            "viet":      str(viet      or ""),
            "english":   str(english   or ""),
            "hanzi":     str(hanzi     or ""),
            "cantonese": str(cantonese or ""),
            "cat":       str(cat       or "General"),
            "notes":     str(notes     or ""),
        })
        dedup.add(normalise(viet))
        cats.add(str(cat or "General"))
    wb.close()
    return rows, dedup, sorted(cats)


def load_progress() -> dict:
    if PROGRESS_FILE.exists():
        with open(PROGRESS_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {"accepted": [], "next_num": 1963}


def save_progress(progress: dict):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def validate_word(w: dict, cats_set: set):
    for field in ("part", "viet", "english", "cat"):
        if not str(w.get(field, "")).strip():
            return False, f"missing '{field}'"
    if w["cat"] not in cats_set:
        return False, f"unknown cat {w['cat']!r}"
    return True, None


def call_api(client, count: int, next_num: int, cats: list, dedup_sample: set) -> list:
    sample = list(dedup_sample)[:60]
    prompt = (
        f"Generate exactly {count} Vietnamese vocabulary words at B2 level (upper-intermediate).\n\n"
        "Return ONLY a JSON array. Each element must have these exact keys:\n"
        f'- "num": integer, starting from {next_num}, sequential\n'
        '- "part": one of: noun, verb, adjective, adverb, phrase, conjunction, preposition\n'
        '- "viet": Vietnamese word or phrase with all correct tones and diacritics\n'
        '- "english": English translation\n'
        '- "hanzi": Chinese characters if the word has a Sino-Vietnamese origin (empty string if not applicable)\n'
        '- "cantonese": Cantonese romanisation matching the hanzi (empty string if hanzi is empty)\n'
        f'- "cat": one of these categories: {", ".join(cats)}\n'
        '- "notes": brief usage note, grammatical note, or mnemonic (empty string if none)\n\n'
        "Requirements:\n"
        "- B2 level vocabulary: abstract concepts (tự do, trách nhiệm), formal/academic verbs,\n"
        "  professional vocabulary (medical, legal, business, technology), complex emotions,\n"
        "  connectives and discourse markers, idiomatic phrases\n"
        "- Vietnamese must include all tonal marks — never omit diacritics\n"
        f"- Do NOT repeat any of these existing words: {', '.join(sample)}\n"
        "- Spread words across all provided categories roughly evenly\n"
        "- Return valid JSON only. No markdown code fences, no commentary outside the array."
    )
    response = client.messages.create(
        model="claude-opus-4-6",
        max_tokens=8192,
        messages=[{"role": "user", "content": prompt}],
    )
    text = response.content[0].text.strip()
    text = re.sub(r'^```(?:json)?\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    return json.loads(text)


def generate_batch(client, target_count: int, next_num: int,
                   cats: list, cats_set: set, dedup: set) -> list:
    """Request words, validate, dedup. Retry for parse failures; accept shortfall for pure dups.
    API/network errors (RateLimitError, APIConnectionError) are NOT caught here — they propagate
    to main() which handles save+exit logic."""
    accepted = []
    remaining = target_count
    parse_retries = 0

    while remaining > 0 and parse_retries <= MAX_RETRIES:
        try:
            words = call_api(client, remaining, next_num + len(accepted), cats, dedup)
        except (json.JSONDecodeError, ValueError) as e:
            # Only catch parse/decode errors here; let API errors bubble up
            parse_retries += 1
            print(f"  ⚠ Parse error (retry {parse_retries}/{MAX_RETRIES}): {e}")
            time.sleep(2)
            continue

        dup_count = parse_fail_count = 0
        for w in words:
            w["num"] = next_num + len(accepted)
            valid, reason = validate_word(w, cats_set)
            if not valid:
                parse_fail_count += 1
                continue
            norm = normalise(w["viet"])
            if norm in dedup:
                dup_count += 1
                continue
            dedup.add(norm)
            accepted.append(w)
            remaining -= 1
            if remaining == 0:
                break

        if remaining > 0:
            if parse_fail_count > 0 and parse_retries < MAX_RETRIES:
                parse_retries += 1
                print(f"  {parse_fail_count} validation failures → retry {parse_retries}/{MAX_RETRIES}")
            elif dup_count > 0 and parse_fail_count == 0:
                print(f"  ⚠ {dup_count} duplicates, no parse failures → accepting shortfall of {remaining}")
                break
            else:
                break

    return accepted


def _get_retry_after(e: anthropic.RateLimitError, default: int = 10) -> int:
    """Extract Retry-After seconds from a RateLimitError response header."""
    try:
        return int(e.response.headers.get("retry-after", default))
    except Exception:
        return default


def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise SystemExit("Error: ANTHROPIC_API_KEY environment variable not set.")

    client = anthropic.Anthropic(api_key=api_key)

    print("Loading existing vocab…")
    existing_rows, dedup, cats = load_existing()
    cats_set = set(cats)
    print(f"  {len(existing_rows)} words, {len(cats)} categories: {cats}")

    progress = load_progress()
    accepted = progress["accepted"]
    next_num = progress["next_num"]
    for w in accepted:
        dedup.add(normalise(w["viet"]))
    print(f"  Resuming from word {next_num} ({len(accepted)} already generated)")

    while next_num <= TARGET_TOTAL:
        batch_size = min(BATCH_SIZE, TARGET_TOTAL - next_num + 1)
        print(f"\nBatch {next_num}–{next_num + batch_size - 1}…")

        for attempt in range(MAX_RETRIES + 1):
            try:
                batch = generate_batch(client, batch_size, next_num, cats, cats_set, dedup)
                break
            except anthropic.RateLimitError as e:
                wait = _get_retry_after(e)
                if attempt >= MAX_RETRIES:
                    save_progress({"accepted": accepted, "next_num": next_num})
                    raise SystemExit(f"Rate limit exhausted. Progress saved to {PROGRESS_FILE}.")
                print(f"  Rate limit — waiting {wait}s (attempt {attempt + 1}/{MAX_RETRIES})…")
                time.sleep(wait)
            except Exception as e:
                if attempt >= MAX_RETRIES:
                    save_progress({"accepted": accepted, "next_num": next_num})
                    raise SystemExit(f"Network error: {e}. Progress saved to {PROGRESS_FILE}.")
                print(f"  Error: {e} — retrying in 10s…")
                time.sleep(10)

        accepted.extend(batch)
        next_num += len(batch)
        save_progress({"accepted": accepted, "next_num": next_num})
        print(f"  ✓ {len(batch)} accepted (total: {len(accepted)})")

    # Write combined output
    print(f"\nWriting {OUTPUT_FILE}…")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = SHEET_NAME
    ws_out.append(["num", "part", "viet", "english", "hanzi", "cantonese", "cat", "notes"])
    for r in existing_rows + accepted:
        ws_out.append([r["num"], r["part"], r["viet"], r["english"],
                       r["hanzi"], r["cantonese"], r["cat"], r["notes"]])
    wb_out.save(OUTPUT_FILE)

    # Summary
    cat_counts   = Counter(w["cat"] for w in accepted)
    total_rows   = len(existing_rows) + len(accepted)
    shortfall    = TARGET_TOTAL - total_rows
    print(f"\n{'=' * 50}")
    print(f"Total rows written : {total_rows}")
    print(f"New words generated: {len(accepted)}")
    if shortfall:
        print(f"Shortfall (dups)   : {shortfall}")
    print(f"By category        : {dict(cat_counts)}")
    print(f"Output             : {OUTPUT_FILE}")

    PROGRESS_FILE.unlink(missing_ok=True)
    print("✓ Done")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""
app.py  —  Language Anki monorepo
==================================
Serves all 3 language apps from one Flask instance.
Run:  python3 app.py
"""

from pathlib import Path
from flask import Flask, jsonify, render_template
import openpyxl

BASE = Path(__file__).parent
app  = Flask(__name__)

# ── Vocab loaders ─────────────────────────────────────────────────────────────

def _load_bahasa():
    src  = BASE / "indonesian" / "Bahasa_Indonesia_Melayu_2000_Words_FINAL.xlsx"
    xlsm = BASE / "indonesian" / "Bahasa_Vocab.xlsm"
    src  = xlsm if xlsm.exists() else src
    print(f"Loading Bahasa vocab from {src}…")

    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    if "SRS_Data" in wb.sheetnames:
        ws = wb["SRS_Data"]
        cards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cards.append({
                "num":     int(row[0]),
                "indo":    str(row[1] or ""),
                "malay":   str(row[2] or ""),
                "english": str(row[3] or ""),
                "cat":     str(row[4] or "General"),
                "contoh":  str(row[5] or ""),
                "eng_ex":  str(row[6] or ""),
            })
    else:
        sheet_name = "Vocab" if "Vocab" in wb.sheetnames else wb.sheetnames[-1]
        ws = wb[sheet_name]
        cards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = (list(row) + [""] * 8)[:8]
            num, cat, eng, indo, malay, contoh_id, _, eng_ex = row
            if not eng:
                continue
            cards.append({
                "num":     int(num) if num else 0,
                "indo":    str(indo      or ""),
                "malay":   str(malay     or ""),
                "english": str(eng       or ""),
                "cat":     str(cat       or "General"),
                "contoh":  str(contoh_id or ""),
                "eng_ex":  str(eng_ex    or ""),
            })
    wb.close()
    print(f"  Loaded {len(cards)} Bahasa cards")
    return cards


def _load_viet():
    src = BASE / "vietnamese" / "viet_vocab_COMPLETE_3000words.xlsx"
    if not src.exists():
        src = BASE / "vietnamese" / "viet_vocab_COMPLETE_1962words.xlsx"
        print("WARNING: 3000-word file not found, falling back to 1962-word file")
    print(f"Loading Viet vocab from {src}…")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb["📚 All Words (Combined)"]
    cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        num, part, viet, english, hanzi, cantonese, cat, notes = (list(row) + [""] * 8)[:8]
        if not viet:
            continue
        cards.append({
            "num":       int(num),
            "part":      str(part      or ""),
            "viet":      str(viet      or ""),
            "english":   str(english   or ""),
            "hanzi":     str(hanzi     or ""),
            "cantonese": str(cantonese or ""),
            "cat":       str(cat       or "General"),
            "notes":     str(notes     or ""),
        })
    wb.close()
    print(f"  Loaded {len(cards)} Viet cards")
    return cards


def _load_spanish():
    src = BASE / "spanish" / "spanish_vocab_3000words.xlsx"
    if not src.exists():
        print(f"WARNING: {src} not found — Spanish will return empty vocab")
        return []
    print(f"Loading Spanish vocab from {src}…")
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    ws = wb.active
    cards = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = (list(row) + [""] * 8)[:8]
        num, cat, english, spanish, gender, notes, example_es, example_en = row
        if not spanish:
            continue
        cards.append({
            "num":        int(num) if num else 0,
            "cat":        str(cat        or "General"),
            "english":    str(english    or ""),
            "spanish":    str(spanish    or ""),
            "gender":     str(gender     or "-"),
            "notes":      str(notes      or ""),
            "example_es": str(example_es or ""),
            "example_en": str(example_en or ""),
        })
    wb.close()
    print(f"  Loaded {len(cards)} Spanish cards")
    return cards


# ── Cache vocab at startup ────────────────────────────────────────────────────
BAHASA_CARDS  = _load_bahasa()
VIET_CARDS    = _load_viet()
SPANISH_CARDS = _load_spanish()

# ── Vocab API routes ──────────────────────────────────────────────────────────
@app.route("/api/vocab/bahasa")
def api_bahasa():
    return jsonify(BAHASA_CARDS)

@app.route("/api/vocab/viet")
def api_viet():
    return jsonify(VIET_CARDS)

@app.route("/api/vocab/spanish")
def api_spanish():
    return jsonify(SPANISH_CARDS)

# ── Page routes ───────────────────────────────────────────────────────────────
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/bahasa")
def bahasa():
    return render_template("bahasa.html")

@app.route("/viet")
def viet():
    return render_template("viet.html")

@app.route("/spanish")
def spanish():
    return render_template("spanish.html")

# ── Run ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)

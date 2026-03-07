#!/usr/bin/env python3
"""
bahasa_anki.py  —  Bahasa Indonesia / Melayu Spaced-Repetition Web App
=======================================================================
Run:  python3 bahasa_anki.py
Then open:  http://localhost:5000
"""

from pathlib import Path
from flask import Flask, jsonify
import openpyxl

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE          = Path(__file__).parent
XLSM          = BASE / "Bahasa_Vocab.xlsm"
XLSX          = BASE / "Bahasa_Indonesia_Melayu_2000_Words_FINAL.xlsx"
SRC           = XLSM if XLSM.exists() else XLSX

app = Flask(__name__)

# ── Load vocab ────────────────────────────────────────────────────────────────
def load_vocab():
    print(f"Loading vocab from {SRC}…")
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    if "SRS_Data" in wb.sheetnames:
        ws = wb["SRS_Data"]
        cards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            cards.append({
                "num":     int(row[0]),
                "indo":    str(row[1] or ""),
                "malay":   str(row[2] or ""),
                "english": str(row[3] or ""),
                "cat":     str(row[4] or "General"),
                "contoh":  str(row[5] or ""),
                "eng_ex":  str(row[6] or ""),
            })
    else:
        sheet_name = "Vocab" if "Vocab" in wb.sheetnames else "Top 2000 Words"
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.sheetnames[-1]
        ws = wb[sheet_name]
        cards = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            row = list(row)[:8]
            if len(row) < 4 or not row[0]:
                continue
            num, cat, eng, indo, malay, contoh_id, contoh_ml, eng_ex = (row + [""] * 8)[:8]
            if not eng:
                continue
            cards.append({
                "num":     int(num) if num else 0,
                "indo":    str(indo or ""),
                "malay":   str(malay or ""),
                "english": str(eng or ""),
                "cat":     str(cat or "General"),
                "contoh":  str(contoh_id or ""),
                "eng_ex":  str(eng_ex or ""),
            })

    wb.close()
    print(f"  Loaded {len(cards)} cards")
    return cards


# ── Cache vocab at startup ────────────────────────────────────────────────────
ALL_CARDS  = load_vocab()

# ── API Routes ────────────────────────────────────────────────────────────────
@app.route("/api/vocab")
def api_vocab():
    return jsonify(ALL_CARDS)

# ── Serve the SPA ─────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return HTML_PAGE

# ── HTML / CSS / JS (single-file SPA) ────────────────────────────────────────
HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Bahasa Anki — Spaced Repetition</title>
<style>
  :root {
    --navy:   #1A3C5E;
    --blue:   #2E6DA4;
    --orange: #E8732A;
    --white:  #FFFFFF;
    --lgrey:  #F0F4F8;
    --dgrey:  #6C7A8A;
    --green:  #27AE60;
    --amber:  #E67E22;
    --red:    #C0392B;
    --teal:   #1ABC9C;
    --card:   #FAFCFF;
    --shadow: #D0DFF0;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
         background: var(--lgrey); min-height: 100vh; }

  /* ── Top bar ── */
  #topbar {
    background: var(--navy); color: var(--white);
    display: flex; align-items: center; padding: 10px 20px; gap: 12px;
    flex-wrap: wrap;
  }
  #topbar h1 { font-size: 18px; font-weight: 700; white-space: nowrap; }
  .mode-btn {
    background: transparent; border: 2px solid rgba(255,255,255,0.3);
    color: rgba(255,255,255,0.6); border-radius: 6px;
    padding: 5px 12px; cursor: pointer; font-size: 13px; font-weight: 600;
    transition: all 0.2s;
  }
  .mode-btn.active { background: var(--blue); border-color: var(--blue); color: white; }
  #cat-select {
    background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.3);
    color: white; padding: 5px 10px; border-radius: 6px; font-size: 13px;
    cursor: pointer;
  }
  #cat-select option { background: var(--navy); }
  .spacer { flex: 1; }
  #stats-btn {
    background: rgba(255,255,255,0.1); border: 1px solid rgba(255,255,255,0.3);
    color: white; padding: 5px 14px; border-radius: 6px; cursor: pointer;
    font-size: 13px; transition: all 0.2s;
  }
  #stats-btn:hover { background: rgba(255,255,255,0.2); }

  /* ── Progress strip ── */
  #progstrip {
    background: #162E48; padding: 6px 20px;
    display: flex; align-items: center; gap: 12px;
  }
  #prog-label { color: rgba(255,255,255,0.7); font-size: 12px; }
  #prog-bar-wrap {
    flex: 1; max-width: 200px; height: 6px;
    background: rgba(255,255,255,0.15); border-radius: 3px; overflow: hidden;
  }
  #prog-bar { height: 100%; background: var(--teal); border-radius: 3px; transition: width 0.4s; width: 0%; }
  #streak { color: #FFD700; font-size: 13px; font-weight: 700; }

  /* ── Main area ── */
  #main { max-width: 720px; margin: 24px auto; padding: 0 16px; }

  /* ── Welcome ── */
  #welcome { text-align: center; padding: 40px 20px; }
  #welcome h2 { color: var(--navy); font-size: 28px; margin-bottom: 12px; }
  #welcome p  { color: var(--dgrey); font-size: 15px; line-height: 1.6; }
  .stat-grid {
    display: grid; grid-template-columns: repeat(4, 1fr);
    gap: 12px; margin: 24px 0;
  }
  .stat-box {
    background: white; border-radius: 10px; padding: 16px 12px; text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
  }
  .stat-box .num { font-size: 26px; font-weight: 700; color: var(--navy); }
  .stat-box .lbl { font-size: 11px; color: var(--dgrey); margin-top: 2px; text-transform: uppercase; }
  .stat-box.due  .num { color: var(--orange); }
  .stat-box.new  .num { color: var(--blue); }
  .stat-box.mat  .num { color: var(--green); }

  /* ── Card ── */
  #card-wrap { display: none; }
  .card {
    background: var(--card); border-radius: 16px;
    box-shadow: 0 4px 20px rgba(26,60,94,0.12);
    overflow: hidden;
  }
  .card-header {
    background: var(--navy); padding: 10px 20px;
    display: flex; align-items: center; justify-content: space-between;
  }
  .card-cat { color: rgba(255,255,255,0.7); font-size: 12px; font-weight: 600;
              text-transform: uppercase; letter-spacing: 0.5px; }
  .card-badge {
    font-size: 11px; padding: 3px 10px; border-radius: 12px;
    font-weight: 600;
  }
  .badge-new      { background: #2E6DA4; color: white; }
  .badge-learning { background: var(--amber); color: white; }
  .badge-review   { background: #8E44AD; color: white; }
  .badge-mature   { background: var(--green); color: white; }
  .card-body { padding: 28px 32px 20px; }
  .card-direction { color: var(--blue); font-size: 13px; font-weight: 600; margin-bottom: 12px; }
  .card-question {
    font-size: 42px; font-weight: 700; color: var(--navy);
    line-height: 1.2; margin-bottom: 8px; word-break: break-word;
  }
  .card-divider { height: 2px; background: var(--lgrey); border-radius: 1px; margin: 16px 0; }

  /* ── Answer ── */
  #answer-area { display: none; }
  .answer-word { font-size: 26px; font-weight: 700; color: var(--green); margin-bottom: 4px; }
  .malay-word  { font-size: 16px; color: #2E7A5E; font-weight: 600; margin-bottom: 12px; }
  .example-id  { font-size: 13px; color: #334455; font-style: italic;
                 background: #F0F8F4; border-left: 3px solid var(--green);
                 padding: 8px 12px; border-radius: 0 6px 6px 0; margin-bottom: 8px; }
  .example-en  { font-size: 12px; color: var(--dgrey); font-style: italic; padding-left: 4px; }

  /* ── Pronounce ── */
  #btn-pronounce {
    margin-top: 12px; background: #D6EAF8; border: none; color: var(--navy);
    padding: 7px 18px; border-radius: 8px; cursor: pointer; font-size: 13px;
    font-weight: 600; transition: background 0.2s;
  }
  #btn-pronounce:hover { background: #BDD4EB; }

  /* ── Show answer / rating ── */
  #action-area { margin-top: 16px; }
  #btn-show {
    width: 100%; background: var(--orange); color: white; border: none;
    padding: 16px; border-radius: 12px; font-size: 16px; font-weight: 700;
    cursor: pointer; transition: background 0.2s; letter-spacing: 0.3px;
  }
  #btn-show:hover { background: #D4621F; }
  #rating-area { display: none; }
  .rating-row { display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; }
  .rating-btn {
    border: none; color: white; padding: 12px 8px; border-radius: 10px;
    cursor: pointer; font-size: 13px; font-weight: 700; line-height: 1.4;
    transition: filter 0.2s; text-align: center;
  }
  .rating-btn:hover { filter: brightness(1.1); }
  .rating-btn:active { filter: brightness(0.9); }
  .btn-again { background: var(--red); }
  .btn-hard  { background: var(--amber); }
  .btn-good  { background: var(--green); }
  .btn-easy  { background: var(--blue); }
  .hint-row  { text-align: center; color: var(--dgrey); font-size: 11px; margin-top: 8px; }

  /* ── Bottom controls ── */
  #bottom-bar {
    display: flex; gap: 10px; margin-top: 16px;
  }
  #btn-start {
    flex: 1; background: var(--teal); color: white; border: none;
    padding: 13px; border-radius: 10px; font-size: 14px; font-weight: 700;
    cursor: pointer; transition: background 0.2s;
  }
  #btn-start:hover { background: #16A085; }
  #btn-quit {
    background: var(--dgrey); color: white; border: none;
    padding: 13px 20px; border-radius: 10px; font-size: 13px; font-weight: 600;
    cursor: pointer;
  }

  /* ── Session end ── */
  #session-end { display: none; text-align: center; padding: 32px 20px; }
  #session-end h2 { color: var(--green); font-size: 32px; margin-bottom: 8px; }
  #session-end .sub { color: var(--dgrey); font-size: 15px; }
  .end-stats { display: flex; justify-content: center; gap: 24px; margin: 20px 0; }
  .end-stat .n { font-size: 28px; font-weight: 700; }
  .end-stat .l { font-size: 12px; color: var(--dgrey); }

  /* ── Stats modal ── */
  #modal-overlay {
    display: none; position: fixed; inset: 0;
    background: rgba(0,0,0,0.5); z-index: 100;
    align-items: center; justify-content: center;
  }
  #modal-overlay.open { display: flex; }
  .modal {
    background: white; border-radius: 16px; width: 560px; max-width: 95vw;
    max-height: 85vh; overflow: hidden; display: flex; flex-direction: column;
    box-shadow: 0 20px 60px rgba(0,0,0,0.3);
  }
  .modal-header {
    background: var(--navy); color: white; padding: 16px 20px;
    display: flex; align-items: center; justify-content: space-between;
  }
  .modal-header h3 { font-size: 16px; }
  .modal-close { background: none; border: none; color: white; font-size: 20px;
                 cursor: pointer; padding: 0 4px; }
  .modal-body { padding: 16px; overflow-y: auto; }
  .overview-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 16px; }
  .ov-box { background: var(--lgrey); border-radius: 8px; padding: 12px; text-align: center; }
  .ov-box .n { font-size: 22px; font-weight: 700; color: var(--navy); }
  .ov-box .l { font-size: 11px; color: var(--dgrey); }
  .cat-table { width: 100%; border-collapse: collapse; font-size: 13px; }
  .cat-table th { background: var(--navy); color: white; padding: 7px 10px; text-align: left; }
  .cat-table td { padding: 6px 10px; border-bottom: 1px solid #EEF2F6; }
  .cat-table tr:nth-child(even) td { background: #F8FAFC; }
  .pct-bar { height: 6px; background: #E0E8F0; border-radius: 3px; overflow: hidden; margin-top: 3px; }
  .pct-fill { height: 100%; background: var(--green); border-radius: 3px; }

  /* ── Keyboard hint ── */
  .kbd { display: inline-block; background: #EEF2F6; border: 1px solid #C8D4E0;
         border-radius: 4px; padding: 1px 6px; font-size: 11px; font-family: monospace;
         color: var(--dgrey); }
</style>
</head>
<body>

<!-- Top bar -->
<div id="topbar">
  <h1>🃏 BAHASA ANKI</h1>
  <button class="mode-btn active" id="btn-id-en" onclick="setMode('id_en')">🇮🇩 ID → EN</button>
  <button class="mode-btn"        id="btn-en-id" onclick="setMode('en_id')">🇬🇧 EN → ID</button>
  <label style="color:rgba(255,255,255,0.6);font-size:13px;">Category:</label>
  <select id="cat-select"></select>
  <div class="spacer"></div>
  <button id="stats-btn" onclick="openStats()">📊 Stats</button>
</div>

<!-- Progress strip -->
<div id="progstrip">
  <span id="prog-label">Ready</span>
  <div id="prog-bar-wrap"><div id="prog-bar"></div></div>
  <span id="streak"></span>
</div>

<!-- Main content -->
<div id="main">

  <!-- Welcome / home screen -->
  <div id="welcome">
    <h2>🇮🇩 Bahasa Indonesia / Melayu</h2>
    <p>Spaced-repetition flashcards — 2 000 words, SM-2 algorithm</p>
    <div class="stat-grid">
      <div class="stat-box due">
        <div class="num" id="ov-due">…</div>
        <div class="lbl">Due Today</div>
      </div>
      <div class="stat-box new">
        <div class="num" id="ov-new">…</div>
        <div class="lbl">New</div>
      </div>
      <div class="stat-box">
        <div class="num" id="ov-learn">…</div>
        <div class="lbl">Learning</div>
      </div>
      <div class="stat-box mat">
        <div class="num" id="ov-mature">…</div>
        <div class="lbl">Mature</div>
      </div>
    </div>
    <p style="color:var(--dgrey);font-size:13px;">
      Keyboard shortcuts: <span class="kbd">Space</span> show answer &nbsp;
      <span class="kbd">1</span>–<span class="kbd">4</span> rate &nbsp;
      <span class="kbd">P</span> pronounce &nbsp;
      <span class="kbd">Enter</span> start
    </p>
  </div>

  <!-- Card view -->
  <div id="card-wrap">
    <div class="card">
      <div class="card-header">
        <span class="card-cat" id="card-cat">—</span>
        <span class="card-badge" id="card-badge">New</span>
      </div>
      <div class="card-body">
        <div class="card-direction" id="card-dir">Indonesian → English</div>
        <div class="card-question"  id="card-q">…</div>
        <div class="card-divider"></div>

        <!-- Answer (hidden until revealed) -->
        <div id="answer-area">
          <div class="answer-word" id="ans-word"></div>
          <div class="malay-word"  id="ans-malay"></div>
          <div class="example-id"  id="ans-contoh" style="display:none"></div>
          <div class="example-en"  id="ans-eng-ex" style="display:none"></div>
        </div>

        <button id="btn-pronounce" onclick="pronounce()">🔊 Pronounce</button>
      </div>
    </div>

    <!-- Action area -->
    <div id="action-area">
      <button id="btn-show" onclick="showAnswer()">👁 Show Answer</button>
      <div id="rating-area">
        <div class="rating-row">
          <button class="rating-btn btn-again" onclick="rate(1)">😰 Again<br><small>forgot</small></button>
          <button class="rating-btn btn-hard"  onclick="rate(2)">😕 Hard<br><small>struggled</small></button>
          <button class="rating-btn btn-good"  onclick="rate(3)">🙂 Good<br><small>knew it</small></button>
          <button class="rating-btn btn-easy"  onclick="rate(4)">😄 Easy<br><small>instant!</small></button>
        </div>
        <div class="hint-row">
          <span class="kbd">1</span> Again &nbsp;
          <span class="kbd">2</span> Hard &nbsp;
          <span class="kbd">3</span> Good &nbsp;
          <span class="kbd">4</span> Easy
        </div>
      </div>
    </div>

  </div><!-- #card-wrap -->

  <!-- Session end -->
  <div id="session-end">
    <h2>🎉 Session Done!</h2>
    <div class="end-stats">
      <div class="end-stat"><div class="n" id="end-cor" style="color:var(--green)">—</div><div class="l">Correct</div></div>
      <div class="end-stat"><div class="n" id="end-wrg" style="color:var(--red)">—</div><div class="l">Wrong</div></div>
      <div class="end-stat"><div class="n" id="end-acc" style="color:var(--navy)">—</div><div class="l">Accuracy</div></div>
    </div>
    <p class="sub" id="end-sub"></p>
  </div>

  <!-- Bottom bar -->
  <div id="bottom-bar">
    <button id="btn-start" onclick="startSession()">▶ Start Session</button>
    <button id="btn-quit"  onclick="goHome()">🏠 Home</button>
  </div>

</div><!-- #main -->

<!-- Stats modal -->
<div id="modal-overlay" onclick="if(event.target===this)closeStats()">
  <div class="modal">
    <div class="modal-header">
      <h3>📊 Your Progress</h3>
      <button class="modal-close" onclick="closeStats()">✕</button>
    </div>
    <div class="modal-body" id="modal-body">Loading…</div>
  </div>
</div>

<script>
// ── Config ─────────────────────────────────────────────────────────────────
const STORAGE_KEY = "bahasa_progress";
let mode        = "id_en";
let allCards    = [];
let queue       = [];
let qpos        = 0;
let curCard     = null;
let answerShown = false;
let sessCorr    = 0;
let sessWrong   = 0;

// ── localStorage ───────────────────────────────────────────────────────────
function loadProgress() {
  try { return JSON.parse(localStorage.getItem(STORAGE_KEY) || "{}"); }
  catch { return {}; }
}
function saveProgress(p) {
  try { localStorage.setItem(STORAGE_KEY, JSON.stringify(p)); }
  catch(e) { console.warn("Could not save progress:", e); }
}

// ── SM-2 ───────────────────────────────────────────────────────────────────
function sm2Update(pd, rating) {
  let interval = pd.interval || 0;
  let ef       = pd.ef       || 2.5;
  let reps     = pd.reps     || 0;
  const grade  = {1:1, 2:3, 3:4, 4:5}[rating];
  if (grade < 3) { reps = 0; interval = 1; }
  else {
    if (reps === 0)      interval = 1;
    else if (reps === 1) interval = 6;
    else                 interval = Math.round(interval * ef);
    reps++;
  }
  ef = ef + (0.1 - (5 - grade) * (0.08 + (5 - grade) * 0.02));
  ef = Math.max(1.3, Math.min(5.0, ef));
  return {
    ...pd,
    interval,
    ef: Math.round(ef * 100) / 100,
    reps,
    next_review: new Date(Date.now() + interval * 86400000).toISOString(),
    correct: (pd.correct || 0) + (grade >= 3 ? 1 : 0),
    wrong:   (pd.wrong   || 0) + (grade <  3 ? 1 : 0),
  };
}
function isDue(pd) {
  if (!pd || !pd.next_review) return true;
  return new Date(pd.next_review) <= new Date();
}
function cardStage(pd) {
  const intv = pd ? (pd.interval || 0) : 0;
  if (intv === 0)  return "new";
  if (intv < 7)    return "learning";
  if (intv < 21)   return "review";
  return "mature";
}

// ── TTS (Web Speech API) ────────────────────────────────────────────────────
function callPronounce(text, lang) {
  if (!window.speechSynthesis) return;
  const clean = text.split("(")[0].replace(/['"\/]/g, "").trim();
  if (!clean) return;
  const utt = new SpeechSynthesisUtterance(clean);
  utt.lang = lang === "id" ? "id-ID" : "en-US";
  speechSynthesis.cancel();
  speechSynthesis.speak(utt);
}

// ── Overview (client-side) ─────────────────────────────────────────────────
function computeOverview() {
  const p = loadProgress();
  let newC = 0, lrn = 0, review = 0, mature = 0, due = 0, cor = 0, wrg = 0;
  for (const c of allCards) {
    const pd = p[String(c.num)] || {};
    const stage = cardStage(pd);
    if (stage === "new")           newC++;
    else if (stage === "learning") lrn++;
    else if (stage === "review")   review++;
    else                           mature++;
    if (isDue(pd)) due++;
    cor += pd.correct || 0;
    wrg += pd.wrong   || 0;
  }
  return { total: allCards.length, new: newC, learning: lrn, review, mature, due,
           correct: cor, wrong: wrg,
           accuracy: (cor + wrg) ? Math.round(100 * cor / (cor + wrg)) : 0 };
}

function refreshOverview() {
  const ov = computeOverview();
  document.getElementById("ov-due").textContent    = ov.due;
  document.getElementById("ov-new").textContent    = ov.new;
  document.getElementById("ov-learn").textContent  = ov.learning + ov.review;
  document.getElementById("ov-mature").textContent = ov.mature;
}

// ── Category stats (client-side) ──────────────────────────────────────────
function computeCatStats() {
  const p = loadProgress();
  const cats = {};
  for (const c of allCards) {
    const pd    = p[String(c.num)] || {};
    const cat   = c.cat;
    const stage = cardStage(pd);
    if (!cats[cat]) cats[cat] = { total:0, new:0, learning:0, review:0, mature:0, due:0 };
    cats[cat].total++;
    cats[cat][stage]++;
    if (isDue(pd)) cats[cat].due++;
  }
  return cats;
}

// ── Init ───────────────────────────────────────────────────────────────────
async function init() {
  allCards = await fetch("/api/vocab").then(r => r.json());
  const cats = [...new Set(allCards.map(c => c.cat))].sort();
  const sel = document.getElementById("cat-select");
  sel.innerHTML = '<option value="All">All categories</option>' +
    cats.map(c => `<option>${c}</option>`).join("");
  refreshOverview();
  showScreen("home");
}

// ── Screen helpers ─────────────────────────────────────────────────────────
function showScreen(name) {
  document.getElementById("welcome").style.display     = name === "home" ? "block" : "none";
  document.getElementById("card-wrap").style.display   = name === "card" ? "block" : "none";
  document.getElementById("session-end").style.display = name === "end"  ? "block" : "none";
}
function goHome() { showScreen("home"); refreshOverview(); setProg(0, "Ready", 0, 0); }

// ── Session ────────────────────────────────────────────────────────────────
function startSession() {
  const cat = document.getElementById("cat-select").value;
  const p   = loadProgress();
  const candidates = allCards.filter(c =>
    (cat === "All" || c.cat === cat) && isDue(p[String(c.num)] || {})
  );
  for (let i = candidates.length - 1; i > 0; i--) {
    const j = Math.floor(Math.random() * (i + 1));
    [candidates[i], candidates[j]] = [candidates[j], candidates[i]];
  }
  queue     = candidates.map(c => c.num);
  qpos      = 0;
  sessCorr  = 0;
  sessWrong = 0;
  if (queue.length === 0) {
    alert("All caught up! No cards due right now.\nCheck back later or choose a different category.");
    return;
  }
  showScreen("card");
  nextCard();
}

function nextCard() {
  if (qpos >= queue.length) { sessionEnd(); return; }
  const num  = queue[qpos];
  const card = allCards.find(c => c.num === num);
  const p    = loadProgress();
  const pd   = p[String(num)] || {};
  curCard     = { ...card, stage: cardStage(pd), interval: pd.interval || 0,
                  ef: pd.ef || 2.5, reps: pd.reps || 0,
                  correct: pd.correct || 0, wrong: pd.wrong || 0 };
  answerShown = false;

  document.getElementById("card-cat").textContent = card.cat;
  const badgeMap = { new:"badge-new", learning:"badge-learning", review:"badge-review", mature:"badge-mature" };
  const badgeTxt = { new:"New", learning:"Learning", review:"Review", mature:"Mature" };
  const badge = document.getElementById("card-badge");
  badge.className   = "card-badge " + (badgeMap[curCard.stage] || "badge-new");
  badge.textContent = (badgeTxt[curCard.stage] || "New") +
    (curCard.interval > 0 ? `  (${curCard.interval}d)` : "");

  if (mode === "id_en") {
    document.getElementById("card-dir").textContent = "Indonesian  \u2192  English";
    document.getElementById("card-q").textContent   = card.indo;
    callPronounce(card.indo, "id");
  } else {
    document.getElementById("card-dir").textContent = "English  \u2192  Indonesian + Malay";
    document.getElementById("card-q").textContent   = card.english;
  }

  document.getElementById("answer-area").style.display  = "none";
  document.getElementById("btn-show").style.display      = "block";
  document.getElementById("rating-area").style.display   = "none";
  setProg(qpos / queue.length * 100,
    `Card ${qpos+1}/${queue.length}   Correct: ${sessCorr}   Wrong: ${sessWrong}`,
    sessCorr, sessWrong);
}

function showAnswer() {
  if (!curCard || answerShown) return;
  answerShown = true;
  const c = curCard;
  let ansWord, malay;
  if (mode === "id_en") {
    ansWord = c.english;
    malay   = c.malay ? `Malay: ${c.malay}` : "";
  } else {
    ansWord = c.indo;
    malay   = c.malay ? `Malay: ${c.malay}` : "";
  }
  document.getElementById("ans-word").textContent  = ansWord;
  document.getElementById("ans-malay").textContent = malay;
  const ctEl = document.getElementById("ans-contoh");
  const exEl = document.getElementById("ans-eng-ex");
  if (c.contoh && c.contoh !== "None") {
    ctEl.textContent = "  " + c.contoh; ctEl.style.display = "block";
  } else { ctEl.style.display = "none"; }
  if (c.eng_ex && c.eng_ex !== "None") {
    exEl.textContent = "  " + c.eng_ex; exEl.style.display = "block";
  } else { exEl.style.display = "none"; }
  document.getElementById("answer-area").style.display = "block";
  document.getElementById("btn-show").style.display    = "none";
  document.getElementById("rating-area").style.display = "block";
  if (mode === "en_id") callPronounce(c.indo, "id");
}

function rate(rating) {
  if (!answerShown || !curCard) return;
  const p  = loadProgress();
  const pd = p[String(curCard.num)] || {};
  p[String(curCard.num)] = sm2Update(pd, rating);
  saveProgress(p);
  if (rating === 1) {
    const insertAt = Math.min(qpos + 2 + Math.floor(Math.random() * 3), queue.length);
    queue.splice(insertAt, 0, curCard.num);
    sessWrong++;
  } else if (rating === 2) {
    sessWrong++;
  } else {
    sessCorr++;
  }
  qpos++;
  nextCard();
}

function sessionEnd() {
  showScreen("end");
  const total = sessCorr + sessWrong;
  const acc   = total ? Math.round(100 * sessCorr / total) : 0;
  document.getElementById("end-cor").textContent = sessCorr;
  document.getElementById("end-wrg").textContent = sessWrong;
  document.getElementById("end-acc").textContent = acc + "%";
  document.getElementById("end-sub").textContent = `${total} cards reviewed. Press Start to go again.`;
  setProg(100, `Done! ${acc}% accuracy`, sessCorr, sessWrong);
  refreshOverview();
}

// ── Pronounce ──────────────────────────────────────────────────────────────
function pronounce() {
  if (!curCard) return;
  callPronounce(curCard.indo, "id");
}

// ── Mode ───────────────────────────────────────────────────────────────────
function setMode(m) {
  mode = m;
  document.getElementById("btn-id-en").classList.toggle("active", m === "id_en");
  document.getElementById("btn-en-id").classList.toggle("active", m === "en_id");
}

// ── Progress bar ───────────────────────────────────────────────────────────
function setProg(pct, label, cor, wrg) {
  document.getElementById("prog-bar").style.width    = pct + "%";
  document.getElementById("prog-label").textContent  = label;
  document.getElementById("streak").textContent      = cor > 0 ? `${cor}` : "";
}

// ── Stats modal ────────────────────────────────────────────────────────────
function openStats() {
  document.getElementById("modal-overlay").classList.add("open");
  const ov   = computeOverview();
  const cats = computeCatStats();
  let html = `
    <div class="overview-grid">
      <div class="ov-box"><div class="n" style="color:var(--orange)">${ov.due}</div><div class="l">Due Today</div></div>
      <div class="ov-box"><div class="n">${ov.total}</div><div class="l">Total Cards</div></div>
      <div class="ov-box"><div class="n" style="color:var(--blue)">${ov.new}</div><div class="l">New</div></div>
      <div class="ov-box"><div class="n" style="color:var(--amber)">${ov.learning + ov.review}</div><div class="l">Learning</div></div>
      <div class="ov-box"><div class="n" style="color:var(--green)">${ov.mature}</div><div class="l">Mature</div></div>
      <div class="ov-box"><div class="n">${ov.accuracy}%</div><div class="l">Accuracy</div></div>
    </div>
    <h4 style="color:var(--navy);margin-bottom:8px;">Progress by Category</h4>
    <table class="cat-table">
      <thead><tr><th>Category</th><th>Total</th><th>Mature</th><th>Due</th></tr></thead>
      <tbody>`;
  for (const [cat, s] of Object.entries(cats).sort()) {
    const pct = s.total ? Math.round(100 * s.mature / s.total) : 0;
    html += `<tr>
      <td>${cat}</td><td>${s.total}</td>
      <td><div>${s.mature} (${pct}%)</div>
          <div class="pct-bar"><div class="pct-fill" style="width:${pct}%"></div></div></td>
      <td style="color:${s.due > 0 ? 'var(--orange)' : 'var(--dgrey)'};font-weight:${s.due>0?700:400}">${s.due || '\u2014'}</td>
    </tr>`;
  }
  html += "</tbody></table>";
  document.getElementById("modal-body").innerHTML = html;
}
function closeStats() {
  document.getElementById("modal-overlay").classList.remove("open");
}

// ── Keyboard shortcuts ─────────────────────────────────────────────────────
document.addEventListener("keydown", e => {
  if (e.target.tagName === "SELECT") return;
  if (e.key === " " || e.key === "Spacebar") { e.preventDefault(); showAnswer(); }
  if (e.key === "1") rate(1);
  if (e.key === "2") rate(2);
  if (e.key === "3") rate(3);
  if (e.key === "4") rate(4);
  if (e.key === "p" || e.key === "P") pronounce();
  if (e.key === "Enter" && document.getElementById("welcome").style.display !== "none") startSession();
  if (e.key === "Escape") closeStats();
});

init();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)

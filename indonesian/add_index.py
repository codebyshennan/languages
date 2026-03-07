"""
add_index.py
Adds a formatted "Index & Guide" sheet to the FINAL workbook.
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from collections import Counter, OrderedDict

PATH = '/Users/wongshennan/Documents/personal/languages/indonesian/Bahasa_Indonesia_Melayu_2000_Words_FINAL.xlsx'

# ── Load workbook & count categories ─────────────────────────────────────────
wb = openpyxl.load_workbook(PATH)
ws_vocab = wb.active

cats = Counter()
total_words = 0
for row in ws_vocab.iter_rows(min_row=2, values_only=True):
    num, cat, eng, *_ = row
    if cat and cat != '—':
        cats[cat] += 1
        total_words += 1

# Priority / study order — arranged by how useful they are for a beginner
STUDY_ORDER = OrderedDict([
    ("Pronouns & Determiners",       ("★★★ Essential",  "🟢 Start here")),
    ("Conjunctions & Connectors",    ("★★★ Essential",  "🟢 Start here")),
    ("Prepositions",                 ("★★★ Essential",  "🟢 Start here")),
    ("Numbers & Quantities",         ("★★★ Essential",  "🟢 Start here")),
    ("Time & Frequency",             ("★★★ Essential",  "🟢 Start here")),
    ("Verbs – General",              ("★★★ Essential",  "🟢 Start here")),
    ("Adjectives",                   ("★★★ Essential",  "🟢 Start here")),
    ("Adverbs & Degree Words",       ("★★★ Essential",  "🟢 Start here")),
    ("Modal & Imperative",           ("★★★ Essential",  "🟢 Start here")),
    ("Expressions & Phrases",        ("★★★ Essential",  "🟢 Start here")),
    ("Food & Drink",                 ("★★☆ High",       "🔵 Week 2–3")),
    ("Family & Relationships",       ("★★☆ High",       "🔵 Week 2–3")),
    ("Body & Health",                ("★★☆ High",       "🔵 Week 2–3")),
    ("Places & Geography",           ("★★☆ High",       "🔵 Week 2–3")),
    ("Transport & Travel",           ("★★☆ High",       "🔵 Week 2–3")),
    ("Household & Objects",          ("★★☆ High",       "🔵 Week 2–3")),
    ("Colors",                       ("★★☆ High",       "🔵 Week 2–3")),
    ("Verbs – Movement & Action",    ("★★☆ High",       "🔵 Week 2–3")),
    ("Verbs – Communication & Mind", ("★★☆ High",       "🔵 Week 2–3")),
    ("Emotions & Personality",       ("★★☆ High",       "🔵 Week 2–3")),
    ("Work & Profession",            ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Education & Knowledge",        ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Nature & Environment",         ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Arts, Culture & Religion",     ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Technology & Media",           ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Business & Finance",           ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Government, Law & Society",    ("★☆☆ Intermediate", "🟡 Month 2")),
    ("Abstract Concepts & Values",   ("★☆☆ Intermediate", "🟡 Month 2")),
    ("General",                      ("Mixed",            "📌 Throughout")),
])

# ── Create the Index sheet (insert as first sheet) ────────────────────────────
if "Index & Guide" in wb.sheetnames:
    del wb["Index & Guide"]
ws = wb.create_sheet("Index & Guide", 0)

# ── Colour palette ────────────────────────────────────────────────────────────
C_TITLE_BG   = "1A3C5E"   # deep navy
C_TITLE_FG   = "FFFFFF"
C_H1_BG      = "2E6DA4"   # medium blue
C_H1_FG      = "FFFFFF"
C_H2_BG      = "D6E4F0"   # pale blue
C_H2_FG      = "1A3C5E"
C_ACCENT     = "E8732A"   # orange accent
C_ROW_ODD    = "FFFFFF"
C_ROW_EVEN   = "EBF3FB"
C_ESSENTIAL  = "D4EDDA"   # light green
C_HIGH       = "D6EAF8"   # light blue
C_INTER      = "FEF9E7"   # light yellow
C_GENERAL    = "F8F9FA"   # near-white
C_BORDER     = "B8CCE4"

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, size=11, color="000000", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(sides="all"):
    s = Side(style="thin", color=C_BORDER)
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    b = Border()
    if "l" in sides: b.left   = s
    if "r" in sides: b.right  = s
    if "t" in sides: b.top    = s
    if "b" in sides: b.bottom = s
    return b

def write(ws, row, col, value, bold=False, size=11, fg="000000",
          bg=None, h="left", v="center", wrap=False, italic=False,
          border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = font(bold=bold, size=size, color=fg, italic=italic)
    cell.alignment = align(h=h, v=v, wrap=wrap)
    if bg:
        cell.fill  = fill(bg)
    if border:
        cell.border = border
    return cell

def section_header(ws, row, col, text, span, bg=C_H1_BG, fg=C_TITLE_FG, size=12):
    write(ws, row, col, text, bold=True, size=size, fg=fg, bg=bg,
          h="left", v="center", wrap=False)
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col + span - 1)
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill  = fill(bg)
        ws.cell(row=row, column=c).alignment = align("left", "center")

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {1:4, 2:32, 3:10, 4:8, 5:16, 6:18, 7:70}
for col, w in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = w

# ─────────────────────────────────────────────────────────────────────────────
# ROW 1–3 : MAIN TITLE BANNER
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 10
ws.row_dimensions[2].height = 38
ws.row_dimensions[3].height = 24
ws.row_dimensions[4].height = 10

write(ws, 2, 2, "📚  Bahasa Indonesia & Bahasa Melayu — Vocabulary Index & Learning Guide",
      bold=True, size=18, fg=C_TITLE_FG, bg=C_TITLE_BG, h="left", v="center")
ws.merge_cells("B2:G2")
for c in range(2, 8):
    ws.cell(2, c).fill = fill(C_TITLE_BG)

write(ws, 3, 2,
      f"  {total_words} words  ·  {len(cats)} categories  ·  Vocabulary list + English examples + study guidance",
      bold=False, size=11, fg=C_TITLE_FG, bg=C_TITLE_BG, h="left", v="center")
ws.merge_cells("B3:G3")
for c in range(2, 8):
    ws.cell(3, c).fill = fill(C_TITLE_BG)

# ─────────────────────────────────────────────────────────────────────────────
# ROW 5 : SECTION — CATEGORY BREAKDOWN
# ─────────────────────────────────────────────────────────────────────────────
r = 5
section_header(ws, r, 2, "  📊  Category Breakdown", 6, bg=C_H1_BG)
ws.row_dimensions[r].height = 26

r += 1
ws.row_dimensions[r].height = 20
for col, txt in zip([2,3,4,5,6,7], ["Category", "Words", "%", "Priority", "Study Phase", "What to expect"]):
    write(ws, r, col, txt, bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG,
          h="center", v="center", border=thin_border())

NOTES = {
    "Pronouns & Determiners":      "saya, kamu, dia, ini, itu — used in every sentence",
    "Conjunctions & Connectors":   "dan, tetapi, karena, jika — glue your sentences together",
    "Prepositions":                "di, ke, dari, pada — critical for location and time",
    "Numbers & Quantities":        "satu–sepuluh, ratus, ribu — shopping, dates, times",
    "Time & Frequency":            "hari, minggu, selalu, kadang — talk about routines",
    "Verbs – General":             "makan, pergi, beli, kerja — action words for daily life",
    "Adjectives":                  "baik, besar, cepat — describe everything around you",
    "Adverbs & Degree Words":      "sangat, sudah, belum, juga — nuance and emphasis",
    "Modal & Imperative":          "jangan, tolong, bisa — requests and prohibitions",
    "Expressions & Phrases":       "terima kasih, maaf, permisi — instant social fluency",
    "Food & Drink":                "nasi, air, makan — survive any warung or kopitiám",
    "Family & Relationships":      "ibu, ayah, teman — introduce yourself and family",
    "Body & Health":               "kepala, sakit, dokter — describe pain and visit a clinic",
    "Places & Geography":          "kota, pasar, rumah — navigate and give directions",
    "Transport & Travel":          "bus, kereta, tiket — get around independently",
    "Household & Objects":         "pintu, meja, kunci — home and everyday objects",
    "Colors":                      "merah, biru, hijau — describe clothes and objects",
    "Verbs – Movement & Action":   "pergi, duduk, berlari — movement and physical action",
    "Verbs – Communication & Mind":"bicara, tanya, pikir — talking, asking, knowing",
    "Emotions & Personality":      "senang, sedih, berani — express and understand feelings",
    "Work & Profession":           "kerja, kantor, rapat — office and professional life",
    "Education & Knowledge":       "sekolah, ujian, ilmu — studying and learning",
    "Nature & Environment":        "pohon, hujan, laut — weather, animals, environment",
    "Arts, Culture & Religion":    "musik, budaya, agama — culture and social life",
    "Technology & Media":          "HP, internet, aplikasi — modern digital life",
    "Business & Finance":          "uang, harga, bisnis — money and commerce",
    "Government, Law & Society":   "hukum, pemerintah, rakyat — society and citizenship",
    "Abstract Concepts & Values":  "tujuan, harapan, kejujuran — deeper conversation",
    "General":                     "High-frequency words that span multiple contexts",
}

BG_MAP = {
    "★★★ Essential":   C_ESSENTIAL,
    "★★☆ High":        C_HIGH,
    "★☆☆ Intermediate": C_INTER,
    "Mixed":           C_GENERAL,
}

r += 1
for cat in STUDY_ORDER:
    count = cats.get(cat, 0)
    pct   = f"{100*count/total_words:.1f}%"
    priority, phase = STUDY_ORDER[cat]
    note  = NOTES.get(cat, "")
    bg    = BG_MAP.get(priority, C_ROW_ODD)
    ws.row_dimensions[r].height = 18
    for c in range(2, 8):
        ws.cell(r, c).fill = fill(bg)
        ws.cell(r, c).border = thin_border()
    write(ws, r, 2, cat,      bold=False, size=10, bg=bg, h="left", v="center", border=thin_border())
    write(ws, r, 3, count,    bold=True,  size=10, bg=bg, h="center", border=thin_border())
    write(ws, r, 4, pct,      bold=False, size=10, bg=bg, h="center", border=thin_border())
    write(ws, r, 5, priority, bold=False, size=10, bg=bg, h="center", border=thin_border())
    write(ws, r, 6, phase,    bold=False, size=10, bg=bg, h="center", border=thin_border())
    write(ws, r, 7, note,     bold=False, size=10, bg=bg, h="left",   border=thin_border())
    r += 1

# Totals row
ws.row_dimensions[r].height = 20
for c in range(2, 8):
    ws.cell(r, c).fill = fill(C_H2_BG)
    ws.cell(r, c).border = thin_border()
write(ws, r, 2, "TOTAL",        bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG, h="left",   border=thin_border())
write(ws, r, 3, total_words,    bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG, h="center", border=thin_border())
write(ws, r, 4, "100%",         bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG, h="center", border=thin_border())
write(ws, r, 5, f"{len(cats)} categories", bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG, h="center", border=thin_border())
r += 2

# ─────────────────────────────────────────────────────────────────────────────
# LEGEND
# ─────────────────────────────────────────────────────────────────────────────
section_header(ws, r, 2, "  🔑  Legend", 6, bg="6C757D")
ws.row_dimensions[r].height = 22
r += 1
legend = [
    (C_ESSENTIAL, "★★★ Essential — Learn these first. Unlocks basic conversation in days."),
    (C_HIGH,      "★★☆ High Priority — Learn in weeks 2–3. Covers travel, food, health."),
    (C_INTER,     "★☆☆ Intermediate — Learn in month 2+. For richer, more specific topics."),
    (C_GENERAL,   "Mixed / General — High-frequency words distributed throughout your study."),
]
for bg, txt in legend:
    ws.row_dimensions[r].height = 18
    ws.cell(r, 2).fill = fill(bg)
    ws.cell(r, 2).border = thin_border()
    write(ws, r, 3, txt, bold=False, size=10, h="left", v="center", wrap=False)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    for c in range(3, 8):
        ws.cell(r, c).border = thin_border()
    r += 1

r += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECTION — HOW TO LEARN BAHASA FAST
# ─────────────────────────────────────────────────────────────────────────────
section_header(ws, r, 2, "  🚀  How to Learn Bahasa Fast — 10 Proven Strategies", 6, bg=C_H1_BG)
ws.row_dimensions[r].height = 26
r += 1

TIPS = [
    ("1", "Master the script — it's phonetic!",
     "Bahasa Indonesia and Malay are written in the Latin alphabet and are almost perfectly phonetic. "
     "Every letter is pronounced consistently. Once you learn the 5 vowel sounds (a, i, u, e, o) and "
     "consonant rules (c = 'ch', ng, ny), you can read anything aloud correctly. "
     "This is your biggest headstart — use it. Spend 1–2 days drilling pronunciation."),

    ("2", "Learn the Essential 300 first (green rows above)",
     "Research shows that the most frequent 300 words cover ~65% of everyday speech. "
     "In this list, focus on: Pronouns, Conjunctions, Prepositions, Time words, Core Verbs, "
     "and basic Adjectives. Learn 15–20 words a day using the 'Contoh Indonesia' example "
     "sentences — context beats rote memorization every time."),

    ("3", "Exploit the zero grammar complexity",
     "Bahasa has NO verb conjugation (no 'I go / he goes / they went'). "
     "NO tenses — instead use time words: 'sudah' (already), 'akan' (will), 'sedang' (currently). "
     "NO gendered nouns. NO plural forms needed (reduplication like 'anak-anak' is optional). "
     "This means you can start forming correct sentences from day 3."),

    ("4", "Use the 'me-/ber-/ter-' prefix system",
     "Indonesian verbs use prefixes to show meaning. Core rule: "
     "'makan' (eat) → 'memakan' (to eat/consume), 'kerja' (work) → 'bekerja' (to work). "
     "Spoken Bahasa often drops prefixes entirely — so 'Saya makan nasi' is perfectly natural. "
     "Learn roots first, prefixes second. Don't let affixes slow you down early on."),

    ("5", "Shadow native speech from day 1",
     "Find YouTube channels, podcasts, or Indonesian/Malay TV (try Netflix series like "
     "'Gadis Kretek' for Indonesian or 'Upin & Ipin' for Malay). "
     "Shadowing — listening and repeating simultaneously — trains your ear AND your mouth. "
     "Even 15 minutes daily of shadowing accelerates fluency dramatically. "
     "Use the 'Contoh Indonesia' column to practice your own shadowing from example sentences."),

    ("6", "Use spaced repetition (Anki or similar)",
     "Export this vocabulary list to Anki (free app). Set reviews to: new cards/day = 20, "
     "review cap = 100. The spaced repetition algorithm schedules cards just before you forget them. "
     "After 3 months, you'll know 1,800+ words with minimal daily effort (~20 min/day). "
     "Use the 'English Example' column as your English prompt, the Indonesian word as the answer."),

    ("7", "Speak from day 1 — mistakes are free",
     "Bahasa speakers are among the world's most encouraging — they will be delighted you tried. "
     "Use iTalki or Preply to book 1-hour sessions with Indonesian or Malaysian tutors for as little "
     "as $5–8/hour. Speak even with a 100-word vocabulary. Simple sentences like "
     "'Saya mau belajar Bahasa. Bisa bantu saya?' open every door."),

    ("8", "Learn cognates — hundreds of free words",
     "Indonesian/Malay borrowed heavily from Dutch, English, Arabic, and Sanskrit. "
     "You already know: 'mobil' (car/automobile), 'komputer' (computer), 'bank' (bank), "
     "'hotel' (hotel), 'restoran' (restaurant), 'telepon' (phone), 'polisi' (police), "
     "'universitas' (university). "
     "English loanwords: drop the suffix or guess phonetically. 'Modern', 'global', 'aktif', "
     "'kreatif', 'nasional', 'internasional' — these are already yours."),

    ("9", "Pair Indonesian with Malay — it doubles your practice material",
     "This list includes both Bahasa Indonesia and Bahasa Melayu (Malaysian). "
     "Because the two are ~80% mutually intelligible, every Indonesian word you learn is nearly "
     "transferable to Malay (and vice versa). Use the 'Bahasa Melayu' column and 'Contoh Melayu' "
     "column to spot differences. Travel to Indonesia AND Malaysia/Singapore with one vocabulary."),

    ("10","Set a 90-day milestone: 1,000 words + real conversation",
     "Week 1–2: Pronunciation + Essentials (300 words). "
     "Week 3–6: High-priority categories + daily Anki + 2× speaking sessions/week. "
     "Week 7–12: Intermediate topics + watch Indonesian/Malay TV without subtitles. "
     "Goal by day 90: hold a 10-minute conversation on daily life topics. "
     "With 1,000 words, you are functionally conversational in Bahasa. This list gives you 2,000."),
]

tip_bgs = [C_ROW_EVEN, C_ROW_ODD]
for i, (num, title, body) in enumerate(TIPS):
    ws.row_dimensions[r].height = 18
    bg = tip_bgs[i % 2]
    write(ws, r, 2, f"  {num}.", bold=True, size=11, fg=C_ACCENT, bg=bg, h="left", v="center")
    write(ws, r, 3, title, bold=True, size=11, fg="1A3C5E", bg=bg, h="left", v="center")
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    for c in range(2, 8):
        ws.cell(r, c).fill = fill(bg)
    r += 1
    ws.row_dimensions[r].height = 72
    write(ws, r, 3, body, bold=False, size=10, bg=bg, h="left", v="top", wrap=True)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    for c in range(2, 8):
        ws.cell(r, c).fill = fill(bg)
    r += 1

r += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECTION — BAHASA INDONESIA vs BAHASA MELAYU
# ─────────────────────────────────────────────────────────────────────────────
section_header(ws, r, 2, "  🔍  Bahasa Indonesia vs Bahasa Melayu — Key Differences", 6, bg=C_H1_BG)
ws.row_dimensions[r].height = 26
r += 1

# Sub-header row
ws.row_dimensions[r].height = 22
for c, txt in zip([2,3,4,5,6,7], ["Category", "Bahasa Indonesia", "", "Bahasa Melayu (Malaysia/Singapore)", "", "Notes"]):
    write(ws, r, c, txt, bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG, h="center", v="center", border=thin_border())
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
r += 1

DIFFS = [
    # (category, indonesian, malay, notes)
    ("Mutual intelligibility",
     "Bahasa Indonesia",
     "Bahasa Melayu / Bahasa Malaysia",
     "~80% mutually intelligible in written form. Spoken may feel faster or more informal. "
     "Both use the same Latin-based Rumi script."),

    ("Official status",
     "Official language of Indonesia (280M speakers). Based on a 1928 standardised form of Malay.",
     "Official language of Malaysia, Brunei; co-official in Singapore. Also called 'Bahasa Malaysia'.",
     "Indonesian is the 4th most spoken language by number of speakers worldwide."),

    ("Spelling system",
     "Reformed 1972 (EYD / EBI). 'c' = /ch/, 'j' = /j/, 'y' = /y/.",
     "Same 1972 reform applied jointly. Virtually identical spelling today.",
     "Pre-1972: Indonesian wrote 'tj' (tjinta), Malay wrote 'ch' (chinta). Now both: 'cinta'."),

    ("Common vocabulary differences",
     "mobil (car)  ·  rumah sakit (hospital)  ·  kereta api (train)\n"
     "pesawat (airplane)  ·  sepeda (bicycle)  ·  kamar mandi (bathroom)\n"
     "handuk (towel)  ·  kulkas (fridge)  ·  polisi (police)\n"
     "tomat (tomato)  ·  wortel (carrot)  ·  jagung (corn)",
     "kereta (car)  ·  hospital (hospital)  ·  kereta api / tren (train)\n"
     "kapal terbang (airplane)  ·  basikal (bicycle)  ·  bilik air (bathroom)\n"
     "tuala (towel)  ·  peti sejuk (fridge)  ·  polis (police)\n"
     "tomato (tomato)  ·  lobak merah (carrot)  ·  jagung (corn)",
     "Everyday nouns often differ. Indonesian borrowed from Dutch; Malay borrowed from English. "
     "Use the 'Bahasa Melayu' column in this workbook to spot these pairs."),

    ("Pronouns",
     "Saya (formal I)  ·  Aku (informal I)\n"
     "Kamu (you, informal)  ·  Anda (you, formal)\n"
     "Dia (he/she)  ·  Mereka (they)\n"
     "Kita (we incl.)  ·  Kami (we excl.)",
     "Saya (formal I)  ·  Aku (informal I)\n"
     "Awak / Kamu (you)  ·  Anda (formal)\n"
     "Dia (he/she)  ·  Mereka (they)\n"
     "Kita (we incl.)  ·  Kami (we excl.)",
     "'Awak' is commonly used in Malaysia for 'you' in friendly speech. "
     "'Encik' (Mr), 'Cik' (Ms), 'Puan' (Mrs) are Malay address forms. "
     "Indonesian uses Bapak/Ibu (formal) or Mas/Mbak (Javanese-influenced)."),

    ("Numbers",
     "satu, dua, tiga, empat, lima\n"
     "enam, tujuh, delapan, sembilan, sepuluh\n"
     "sebelas, dua belas … dua puluh\n"
     "seratus, seribu, sejuta",
     "satu, dua, tiga, empat, lima\n"
     "enam, tujuh, lapan, sembilan, sepuluh\n"
     "sebelas, dua belas … dua puluh\n"
     "seratus, seribu, sejuta",
     "Key difference: 'delapan' (ID) vs 'lapan' (MY) for 8. Otherwise nearly identical."),

    ("Verb affixes",
     "me- prefix common in writing: menulis (write), membaca (read), makan (eat — drop-ok in speech). "
     "ber- for states: berjalan (walk), bekerja (work).",
     "Same me-/ber- system. In informal Malaysian speech, prefixes also frequently dropped. "
     "ter- (accidental/superlative) works identically in both.",
     "Both languages share the same affix system (me-, ber-, ter-, pe-, -an, -kan, -i). "
     "Roots are usually shared; differences are mainly vocabulary-level."),

    ("Formality & register",
     "Formal (written/news): full affixes, standard vocabulary.\n"
     "Colloquial Jakarta: 'gue/gw' (I), 'lo' (you), 'nggak/gak' (not), 'udah' (already), "
     "'aja' (just), 'dong', 'sih', 'deh' — particles everywhere.",
     "Formal (written/official): standard Malay.\n"
     "Colloquial KL: 'lah', 'loh', 'mah', 'kan' — sentence-final particles. "
     "'tak' (not), 'nak' (want), 'dah' (already). Singlish/Manglish mixes in English.",
     "Colloquial registers are VERY different. Focus on formal/standard Bahasa first — "
     "it is understood everywhere. Pick up regional colloquialisms through immersion."),

    ("Loanword origins",
     "Dutch influence: polisi (police), gratis (free), kantor (office), "
     "wortel (carrot), handuk (towel), kopling (clutch).",
     "English influence: polis (police), percuma (free), ofis (office), "
     "lobak merah (carrot), tuala (towel), klac (clutch).",
     "Indonesia was colonized by the Dutch; Malaysia/Singapore by the British. "
     "This single fact explains most vocabulary differences. "
     "In technology and modern terms, both borrow heavily from English."),

    ("Arabic loanwords (shared)",
     "waktu (time), ilmu (knowledge), fikir (think), kabar (news), "
     "masjid (mosque), wajib (obligatory), hadir (present).",
     "waktu (time), ilmu (knowledge), fikir (think), khabar (news), "
     "masjid (mosque), wajib (obligatory), hadir (present).",
     "Both languages share a large Arabic vocabulary due to Islamic influence, "
     "especially in religion, law, and formal/academic contexts."),

    ("Useful phrases",
     "Selamat pagi (Good morning)\n"
     "Terima kasih (Thank you) · Sama-sama / Kembali (You're welcome)\n"
     "Maaf / Permisi (Sorry / Excuse me)\n"
     "Tidak apa-apa (It's okay)\n"
     "Saya tidak mengerti (I don't understand)\n"
     "Bisa pelan-pelan? (Can you speak slower?)",
     "Selamat pagi (Good morning)\n"
     "Terima kasih (Thank you) · Sama-sama / Terima kasih kembali (You're welcome)\n"
     "Maaf / Maafkan saya (Sorry / Excuse me)\n"
     "Tidak mengapa / Takpe (It's okay)\n"
     "Saya tidak faham (I don't understand)\n"
     "Boleh cakap perlahan sikit? (Can you speak slower?)",
     "Core social phrases are very similar. Memorise the Indonesian set — "
     "Malaysians will understand you and appreciate the effort. "
     "Pick up Malay colloquialisms later through exposure."),
]

diff_bgs = [C_ROW_EVEN, C_ROW_ODD]
for i, (cat, indo, malay_txt, note) in enumerate(DIFFS):
    bg = diff_bgs[i % 2]
    # estimate rows needed based on longest text
    lines = max(indo.count('\n'), malay_txt.count('\n'), note.count('\n')) + 2
    row_h = max(54, lines * 18)
    ws.row_dimensions[r].height = row_h
    write(ws, r, 2, cat,       bold=True,  size=10, fg="1A3C5E", bg=bg, h="left", v="top", wrap=True, border=thin_border())
    write(ws, r, 3, indo,      bold=False, size=10, bg=bg, h="left", v="top", wrap=True, border=thin_border())
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    write(ws, r, 5, malay_txt, bold=False, size=10, bg=bg, h="left", v="top", wrap=True, border=thin_border())
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    write(ws, r, 7, note,      bold=False, size=10, bg=bg, h="left", v="top", wrap=True, italic=True, border=thin_border())
    for c in range(2, 8):
        ws.cell(r, c).fill = fill(bg)
    r += 1

r += 1

# ─────────────────────────────────────────────────────────────────────────────
# SECTION — QUICK REFERENCE CHEAT SHEET
# ─────────────────────────────────────────────────────────────────────────────
section_header(ws, r, 2, "  ⚡  Quick-Start Cheat Sheet — First 50 Must-Know Words", 6, bg=C_H1_BG)
ws.row_dimensions[r].height = 26
r += 1

CHEAT = [
    # (Indonesian, Malay, English, Example)
    ("saya / aku",    "saya / aku",    "I / me",          "Saya mau belajar. (I want to learn.)"),
    ("kamu / Anda",   "awak / Anda",   "you",             "Kamu dari mana? (Where are you from?)"),
    ("dia",           "dia",           "he / she",        "Dia sangat baik. (He/she is very kind.)"),
    ("kami / kita",   "kami / kita",   "we",              "Kami pergi bersama. (We go together.)"),
    ("ini / itu",     "ini / itu",     "this / that",     "Ini apa? (What is this?)"),
    ("ada",           "ada",           "there is / exist", "Ada restoran di sini? (Is there a restaurant here?)"),
    ("tidak / bukan", "tidak / bukan", "no / not",        "Saya tidak mau. (I don't want to.)"),
    ("ya / iya",      "ya",            "yes",             "Ya, saya mengerti. (Yes, I understand.)"),
    ("dan",           "dan",           "and",             "Nasi dan ayam. (Rice and chicken.)"),
    ("atau",          "atau",          "or",              "Teh atau kopi? (Tea or coffee?)"),
    ("tapi / tetapi", "tetapi",        "but",             "Enak tapi mahal. (Delicious but expensive.)"),
    ("karena",        "sebab / kerana","because",         "Saya lapar karena belum makan."),
    ("mau / ingin",   "nak / mahu",    "want / would like","Saya mau air putih. (I want plain water.)"),
    ("bisa / dapat",  "boleh / dapat", "can / may",       "Bisa tolong? (Can you help?)"),
    ("pergi",         "pergi",         "to go",           "Saya pergi ke pasar. (I'm going to the market.)"),
    ("makan",         "makan",         "to eat",          "Makan dulu! (Let's eat first!)"),
    ("minum",         "minum",         "to drink",        "Mau minum apa? (What would you like to drink?)"),
    ("beli",          "beli",          "to buy",          "Di mana beli tiket? (Where to buy a ticket?)"),
    ("berapa",        "berapa",        "how much / many", "Berapa harganya? (How much does it cost?)"),
    ("di mana",       "di mana",       "where",           "Di mana toilet? (Where is the toilet?)"),
    ("sudah",         "sudah / dah",   "already / done",  "Saya sudah makan. (I have already eaten.)"),
    ("belum",         "belum",         "not yet",         "Belum, nanti ya. (Not yet, later.)"),
    ("akan",          "akan",          "will / going to", "Saya akan datang. (I will come.)"),
    ("sangat",        "sangat / amat", "very",            "Sangat enak! (Very delicious!)"),
    ("tolong",        "tolong",        "please / help",   "Tolong ulangi. (Please repeat.)"),
    ("maaf",          "maaf",          "sorry",           "Maaf, saya terlambat. (Sorry, I'm late.)"),
    ("terima kasih",  "terima kasih",  "thank you",       "Terima kasih banyak! (Thank you very much!)"),
    ("sama-sama",     "sama-sama",     "you're welcome",  "Sama-sama! (You're welcome!)"),
    ("selamat",       "selamat",       "greetings / safe","Selamat pagi! (Good morning!)"),
    ("hari ini",      "hari ini",      "today",           "Hari ini cuacanya bagus. (Today the weather is nice.)"),
]

ws.row_dimensions[r].height = 20
for c, txt in zip([2,3,4,5,7], ["#", "Bahasa Indonesia", "Bahasa Melayu", "English", "Example sentence"]):
    write(ws, r, c, txt, bold=True, size=10, fg=C_H2_FG, bg=C_H2_BG, h="center", v="center", border=thin_border())
r += 1

for i, (indo, malay_w, eng, ex) in enumerate(CHEAT):
    bg = diff_bgs[i % 2]
    ws.row_dimensions[r].height = 18
    write(ws, r, 2, i+1,     bold=True,  size=10, fg="888888", bg=bg, h="center", border=thin_border())
    write(ws, r, 3, indo,    bold=True,  size=10, fg="1A3C5E", bg=bg, h="left",   border=thin_border())
    write(ws, r, 4, malay_w, bold=False, size=10, bg=bg, h="left",   border=thin_border())
    write(ws, r, 5, eng,     bold=True,  size=10, fg=C_ACCENT, bg=bg, h="left",   border=thin_border())
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    write(ws, r, 7, ex,      bold=False, size=10, bg=bg, italic=True, h="left",   border=thin_border())
    for c in range(2, 8):
        ws.cell(r, c).fill = fill(bg)
    r += 1

r += 2

# ─────────────────────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[r].height = 22
write(ws, r, 2,
      "  💡  Tip: Use the 'Vocab' sheet to study. Filter by Category column to focus on one topic at a time. "
      "Sort by # column to restore original frequency order.",
      bold=False, size=10, fg=C_TITLE_FG, bg=C_TITLE_BG, h="left", v="center", wrap=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
for c in range(2, 8):
    ws.cell(r, c).fill = fill(C_TITLE_BG)

# ── Freeze top rows & set zoom ────────────────────────────────────────────────
ws.freeze_panes = "B5"
ws.sheet_view.zoomScale = 95
ws.sheet_properties.tabColor = "2E6DA4"

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(PATH)
print(f"Done. Index sheet added → {PATH}")
print(f"  Rows written: ~{r}")
print(f"  Categories shown: {len(cats)}")
print(f"  Tips: {len(TIPS)}  |  Diff sections: {len(DIFFS)}  |  Cheat sheet rows: {len(CHEAT)}")

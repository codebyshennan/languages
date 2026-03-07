"""
build_final.py
==============
1. Beautifully formats the Vocab sheet (category colors, borders, freeze, filter)
2. Adds 🔊 pronunciation hyperlinks (Google Translate) for every word
3. Creates Anki Flashcard sheet (cell-based UI)
4. Creates hidden SRS_Data sheet (spaced-repetition tracking)
5. Saves as .xlsm
6. Injects full VBA Anki engine via xlwings
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from urllib.parse import quote
from collections import defaultdict
import subprocess, sys, time, os

SRC  = '/Users/wongshennan/Documents/personal/languages/indonesian/Bahasa_Indonesia_Melayu_2000_Words_FINAL.xlsx'
DEST = '/Users/wongshennan/Documents/personal/languages/indonesian/Bahasa_Vocab.xlsm'

# ── CATEGORY COLOUR MAP ───────────────────────────────────────────────────────
CAT_COLORS = {
    "Pronouns & Determiners":       ("C8DEFF", "A8C4F0"),
    "Conjunctions & Connectors":    ("C8EDD0", "A8D8B4"),
    "Prepositions":                 ("FFF4BC", "FFE88A"),
    "Modal & Imperative":           ("FFE0C8", "FFCAA0"),
    "Numbers & Quantities":         ("E8D5FF", "D4B8F8"),
    "Time & Frequency":             ("D0F0FF", "AADCF8"),
    "Colors":                       ("FFD6EC", "FFB8DC"),
    "Verbs – General":              ("B8F0E8", "8EDDD0"),
    "Verbs – Movement & Action":    ("C0EED8", "98DEB8"),
    "Verbs – Communication & Mind": ("D4F4C8", "B4E4A8"),
    "Adjectives":                   ("FFE8B8", "FFD488"),
    "Adverbs & Degree Words":       ("F0D8FF", "E0B8FF"),
    "Expressions & Phrases":        ("E8E8FF", "D0D0FF"),
    "Food & Drink":                 ("FFF0D0", "FFE0A0"),
    "Family & Relationships":       ("D8FFD8", "B0F0B0"),
    "Body & Health":                ("FFD8D8", "FFB8B8"),
    "Nature & Environment":         ("C8F8E8", "A0ECD0"),
    "Places & Geography":           ("D8EEFF", "B4D8FF"),
    "Transport & Travel":           ("FFE8D0", "FFD0A8"),
    "Household & Objects":          ("F0FFD0", "DFFFA8"),
    "Technology & Media":           ("E8D8FF", "D0B8FF"),
    "Education & Knowledge":        ("D0F8FF", "A8EEFF"),
    "Business & Finance":           ("FFFFD0", "FFFF98"),
    "Government, Law & Society":    ("E8E8E8", "D4D4D4"),
    "Work & Profession":            ("F8E8D8", "F0D0B8"),
    "Arts, Culture & Religion":     ("FFD8F0", "FFB8E4"),
    "Abstract Concepts & Values":   ("E0E0E0", "C8C8C8"),
    "Emotions & Personality":       ("FFD8C0", "FFB898"),
    "General":                      ("F8F8F8", "EEEEEE"),
    "—":                            ("FFFFFF", "F0F0F0"),
}

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def font(bold=False, size=10, color="000000", italic=False, name="Calibri"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(style="thin", color="CCCCCC"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

# ── LOAD WORKBOOK ─────────────────────────────────────────────────────────────
print("Loading workbook…")
wb = openpyxl.load_workbook(SRC)

# ── 1. FORMAT VOCAB SHEET ─────────────────────────────────────────────────────
print("Formatting Vocab sheet…")

# Find and rename the vocab sheet
ws = wb["Top 2000 Words"] if "Top 2000 Words" in wb.sheetnames else wb.active
ws.title = "Vocab"

# Column layout (existing): #, Category, English, Bahasa Indonesia, Bahasa Melayu,
#                           Contoh Indonesia, Contoh Melayu, English Example
# We add col 9: 🔊

COLS = {
    1: ("#",               4),
    2: ("Category",       22),
    3: ("English",        26),
    4: ("Bahasa Indonesia",20),
    5: ("Bahasa Melayu",  20),
    6: ("Contoh Indonesia",34),
    7: ("Contoh Melayu",  34),
    8: ("English Example",38),
    9: ("🔊",              5),
}

for col, (_, width) in COLS.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# Header row
ws.row_dimensions[1].height = 26
HEADER_BG = "1A3C5E"
HEADER_FG = "FFFFFF"
for col, (label, _) in COLS.items():
    cell = ws.cell(1, col, label)
    cell.font      = Font(bold=True, size=10, color=HEADER_FG, name="Calibri")
    cell.fill      = fill(HEADER_BG)
    cell.alignment = align("center", "center")
    cell.border    = border("thin", "2E6DA4")

# Data rows — colour by category, alternate shade within each block
last_cat = None
shade_idx = 0

all_rows = list(ws.iter_rows(min_row=2, values_only=True))
total_rows = len(all_rows)

for i, row_data in enumerate(all_rows):
    r = i + 2
    ws.row_dimensions[r].height = 16

    row_data = row_data[:8]  # safety: cap at 8 columns
    num, cat, eng, indo, malay, contoh_id, contoh_ml, eng_ex = row_data

    if cat != last_cat:
        last_cat = cat
        shade_idx = 0
    else:
        shade_idx += 1

    colors = CAT_COLORS.get(cat, ("F8F8F8", "EEEEEE"))
    bg = colors[shade_idx % 2]

    # existing 8 columns
    for col in range(1, 9):
        cell = ws.cell(r, col)
        cell.fill      = fill(bg)
        cell.alignment = align("left", "center", wrap=(col in (6, 7, 8)))
        cell.border    = border("thin", "D8D8D8")
        cell.font      = font(size=10)

    # Style specific columns
    ws.cell(r, 1).alignment = align("center", "center")
    ws.cell(r, 1).font      = font(size=9, color="888888")
    ws.cell(r, 2).font      = font(size=9, color="2E5F8A", bold=True)
    ws.cell(r, 3).font      = font(size=10, bold=True, color="1A3C5E")
    ws.cell(r, 4).font      = font(size=10, color="0A5C2E")
    ws.cell(r, 5).font      = font(size=10, color="3A2A6E")
    ws.cell(r, 6).font      = font(size=9,  color="555555", italic=True)
    ws.cell(r, 7).font      = font(size=9,  color="555555", italic=True)
    ws.cell(r, 8).font      = font(size=9,  color="335577")

    # Left accent border by category (dark version of bg color)
    cat_border = border("medium", "2E6DA4")
    ws.cell(r, 1).border = Border(
        left  = Side(style="medium",  color="2E6DA4"),
        right = Side(style="thin",    color="D8D8D8"),
        top   = Side(style="thin",    color="D8D8D8"),
        bottom= Side(style="thin",    color="D8D8D8"),
    )

    # ── 🔊 Audio column (col 9) ─────────────────────────────────────────────
    cell9 = ws.cell(r, 9)
    cell9.fill      = fill(bg)
    cell9.border    = border("thin", "D8D8D8")
    cell9.alignment = align("center", "center")

    # Generate Google Translate URL for the Indonesian word
    if indo and indo != "— (to be added)":
        word_for_url = str(indo).split('/')[0].strip()
        encoded = quote(word_for_url)
        gt_url = f"https://translate.google.com/?sl=id&tl=en&text={encoded}&op=translate"
        cell9.value     = "🔊"
        cell9.hyperlink = gt_url
        cell9.font      = Font(bold=True, size=11, color="1155CC", name="Calibri")
    else:
        cell9.value = ""

# Freeze header
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(9)}1"
ws.sheet_view.zoomScale = 100
ws.sheet_properties.tabColor = "2E6DA4"
print(f"  ✓ Formatted {total_rows} rows with category colours + 🔊 links")

# ── 2. SRS_DATA SHEET (hidden tracking) ───────────────────────────────────────
print("Creating SRS_Data sheet…")

if "SRS_Data" in wb.sheetnames:
    del wb["SRS_Data"]
ws_srs = wb.create_sheet("SRS_Data")
ws_srs.sheet_state = "hidden"

SRS_HEADERS = ["#","Indonesian","Malay","English","Category",
               "Contoh","EngExample","Interval","EF","NextReview",
               "Reps","Correct","Wrong"]
for c, h in enumerate(SRS_HEADERS, 1):
    ws_srs.cell(1, c, h).font = Font(bold=True, size=10)

srs_row = 2
for row_data in all_rows:
    row_data = row_data[:8]
    num, cat, eng, indo, malay, contoh_id, contoh_ml, eng_ex = row_data
    if not num or str(num) == "—":
        continue
    if not eng or eng == "— (to be added)":
        continue
    ws_srs.cell(srs_row, 1, num)
    ws_srs.cell(srs_row, 2, indo  or "")
    ws_srs.cell(srs_row, 3, malay or "")
    ws_srs.cell(srs_row, 4, eng)
    ws_srs.cell(srs_row, 5, cat  or "General")
    ws_srs.cell(srs_row, 6, contoh_id or "")
    ws_srs.cell(srs_row, 7, eng_ex or "")
    ws_srs.cell(srs_row, 8, 0)    # Interval
    ws_srs.cell(srs_row, 9, 2.5)  # EF
    ws_srs.cell(srs_row, 10, "")  # NextReview
    ws_srs.cell(srs_row, 11, 0)   # Reps
    ws_srs.cell(srs_row, 12, 0)   # Correct
    ws_srs.cell(srs_row, 13, 0)   # Wrong
    srs_row += 1

print(f"  ✓ SRS_Data populated with {srs_row-2} cards")

# ── 3. FLASHCARD SHEET ────────────────────────────────────────────────────────
print("Creating Flashcard sheet…")

if "Flashcard" in wb.sheetnames:
    del wb["Flashcard"]
ws_fc = wb.create_sheet("Flashcard", 1)   # second tab after Index & Guide
ws_fc.sheet_properties.tabColor = "E8732A"
ws_fc.sheet_view.zoomScale = 110
ws_fc.sheet_view.showGridLines = False

# Column widths
fc_widths = {1:2, 2:2, 3:18, 4:36, 5:18, 6:16, 7:18, 8:22, 9:2}
for c, w in fc_widths.items():
    ws_fc.column_dimensions[get_column_letter(c)].width = w

# Row heights
fc_heights = {
    1:8, 2:40, 3:8, 4:26, 5:22, 6:8,
    7:20, 8:58, 9:20, 10:8,
    11:30, 12:24, 13:24, 14:24, 15:8,
    16:22, 17:52, 18:52, 19:52, 20:52, 21:8,
    22:24, 23:8, 24:22, 25:14,
}
for r, h in fc_heights.items():
    ws_fc.row_dimensions[r].height = h

NAVY  = "1A3C5E"
ORANGE= "E8732A"
WHITE = "FFFFFF"
LGREY = "F0F4F8"
DGREY = "888888"
GREEN = "27AE60"
AMBER = "F39C12"
RED   = "C0392B"
BLUE  = "2980B9"
CARD  = "FAFCFF"
CARD2 = "EEF4FB"

def fc_write(r, c, val, bold=False, size=11, fg="000000", bg=None,
             h="center", v="center", wrap=False, italic=False, span_end_col=None):
    cell = ws_fc.cell(r, c, val)
    cell.font      = Font(bold=bold, size=size, color=fg, name="Calibri", italic=italic)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if bg:
        cell.fill = fill(bg)
    if span_end_col:
        ws_fc.merge_cells(start_row=r, start_column=c,
                          end_row=r, end_column=span_end_col)
        for cc in range(c+1, span_end_col+1):
            ws_fc.cell(r, cc).fill = fill(bg or "FFFFFF")

# ── Title banner ──────────────────────────────────────────────────────────────
fc_write(2, 3, "🃏  BAHASA FLASH CARDS  •  Spaced Repetition",
         bold=True, size=15, fg=WHITE, bg=NAVY, h="center", span_end_col=8)

# ── Config row ────────────────────────────────────────────────────────────────
# Mode label + dropdown hint
fc_write(4, 3, "Mode:", bold=True, size=10, fg=NAVY, bg=LGREY, h="right")
fc_write(4, 4, "Indonesian → English",
         bold=False, size=10, fg="333333", bg=WHITE, h="left")  # VBA writes here

fc_write(4, 6, "Category:", bold=True, size=10, fg=NAVY, bg=LGREY, h="right")
fc_write(4, 7, "All", bold=False, size=10, fg="333333", bg=WHITE, h="left")

# Card status badge (VBA updates)
fc_write(5, 3, "⭐ New card", bold=False, size=9, fg=DGREY, bg=LGREY,
         h="center", span_end_col=4)
# Progress counter
fc_write(5, 6, "Card 0 of 0   ✅ 0   ❌ 0",
         bold=False, size=9, fg=DGREY, bg=LGREY, h="center", span_end_col=8)

# ── Card question area ────────────────────────────────────────────────────────
# Top border of card
for c in range(3, 9):
    ws_fc.cell(6, c).fill = fill("D8E8F8")

# Language direction label
fc_write(7, 3, "🇮🇩  Indonesian → English",
         bold=True, size=10, fg=BLUE, bg=CARD, h="center", span_end_col=8)

# THE WORD — big and prominent
fc_write(8, 3, "Press  ▶ START  to begin",
         bold=True, size=28, fg=NAVY, bg=CARD, h="center", v="center",
         span_end_col=8)
ws_fc.row_dimensions[8].height = 72

# Phonetic hint row
fc_write(9, 3, "",
         bold=False, size=12, fg="6688AA", bg=CARD, h="center", span_end_col=8)

# Bottom border of card
for c in range(3, 9):
    ws_fc.cell(10, c).fill = fill("D8E8F8")

# ── Answer area (VBA reveals these) ──────────────────────────────────────────
fc_write(11, 3, "",  bold=True,  size=20, fg=GREEN, bg=CARD2, h="center", span_end_col=8)
fc_write(12, 3, "",  bold=False, size=11, fg="447766", bg=CARD2, h="center",
         italic=True, span_end_col=8)
fc_write(13, 3, "",  bold=False, size=10, fg="556677", bg=CARD2, h="left",
         wrap=True, span_end_col=8)
fc_write(14, 3, "",  bold=False, size=10, fg="778899", bg=CARD2, h="left",
         wrap=True, italic=True, span_end_col=8)

# ── Divider ───────────────────────────────────────────────────────────────────
for c in range(3, 9):
    ws_fc.cell(15, c).fill = fill(NAVY)

# ── Rating buttons area label ─────────────────────────────────────────────────
fc_write(16, 3, "How well did you remember?",
         bold=True, size=10, fg=NAVY, bg=LGREY, h="center", span_end_col=8)

# Button placeholder cells — VBA draws real shapes over these
BTN_CONFIGS = [
    # (row, col, label, bg, fg)
    (17, 3, "😰  Again\n(forgot)",      "C0392B", WHITE),
    (17, 5, "😕  Hard\n(struggled)",    "D35400", WHITE),
    (18, 3, "🙂  Good\n(knew it)",      "27AE60", WHITE),
    (18, 5, "😄  Easy\n(instant!)",     "2980B9", WHITE),
]
for br, bc, blbl, bbg, bfg in BTN_CONFIGS:
    ws_fc.row_dimensions[br].height = 52
    c = ws_fc.cell(br, bc, blbl)
    c.font      = Font(bold=True, size=12, color=bfg, name="Calibri")
    c.fill      = fill(bbg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_fc.merge_cells(start_row=br, start_column=bc,
                      end_row=br, end_column=bc+1)
    ws_fc.cell(br, bc+1).fill = fill(bbg)

# Show Answer button placeholder
ws_fc.row_dimensions[19].height = 52
c = ws_fc.cell(19, 3, "👁  Show Answer")
c.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
c.fill      = fill(ORANGE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_fc.merge_cells("C19:H19")
for cc in range(3, 9):
    ws_fc.cell(19, cc).fill = fill(ORANGE)

# Pronounce button
ws_fc.row_dimensions[20].height = 36
c = ws_fc.cell(20, 3, "🔊  Pronounce (Damayanti voice)")
c.font      = Font(bold=True, size=11, color=NAVY, name="Calibri")
c.fill      = fill("D6EAF8")
c.alignment = Alignment(horizontal="center", vertical="center")
ws_fc.merge_cells("C20:H20")
for cc in range(3, 9):
    ws_fc.cell(20, cc).fill = fill("D6EAF8")

# ── Control row ───────────────────────────────────────────────────────────────
fc_write(22, 3, "▶  START / RESTART",
         bold=True, size=11, fg=WHITE, bg="1ABC9C", h="center", span_end_col=4)
fc_write(22, 5, "📊  Stats",
         bold=True, size=11, fg=WHITE, bg="8E44AD", h="center", span_end_col=6)
fc_write(22, 7, "⚙  Setup",
         bold=True, size=10, fg=WHITE, bg="7F8C8D", h="center", span_end_col=8)
ws_fc.row_dimensions[22].height = 36

# ── Hidden config cells (VBA reads/writes) ────────────────────────────────────
# H3 = mode ("ID_EN" or "EN_ID"), H4 = category
ws_fc.cell(3, 8, "ID_EN").font = Font(color="FFFFFF", size=1)   # invisible
ws_fc.cell(4, 8, "All").font   = Font(color="FFFFFF", size=1)   # invisible

# ── Instructions ──────────────────────────────────────────────────────────────
fc_write(24, 3,
    "Click ▶ START to begin your session.  "
    "Rate each card after seeing the answer.  "
    "'Again' re-queues the card until you get it right.",
    bold=False, size=9, fg=DGREY, bg=LGREY, h="left", wrap=True, span_end_col=8)

# Fill all unused cells with light background
for r in range(1, 30):
    for c in range(1, 10):
        cell = ws_fc.cell(r, c)
        if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
            cell.fill = fill(LGREY)

print("  ✓ Flashcard sheet created")

# ── SAVE AS XLSM ─────────────────────────────────────────────────────────────
# Reorder sheets: Index & Guide → Flashcard → Vocab → SRS_Data
sheet_order = []
for name in ["Index & Guide", "Flashcard", "Vocab", "SRS_Data"]:
    if name in wb.sheetnames:
        sheet_order.append(wb[name])
# Add any remaining sheets
for ws_i in wb.worksheets:
    if ws_i not in sheet_order:
        sheet_order.append(ws_i)
wb._sheets = sheet_order

print(f"Saving to {DEST}…")
wb.save(DEST)
print("  ✓ Saved")

# ── VBA CODE ─────────────────────────────────────────────────────────────────
VBA_CODE = r'''
Option Explicit

' =============================================================================
'  BAHASA INDONESIA / MALAY — SPACED REPETITION FLASHCARD ENGINE
'  SM-2 Algorithm  |  Damayanti (Indonesian) TTS  |  xlsm module
' =============================================================================

' Sheet / column constants
Const VOCAB_SH  As String = "Vocab"
Const SRS_SH    As String = "SRS_Data"
Const CARD_SH   As String = "Flashcard"

Const C_NUM  As Integer = 1
Const C_INDO As Integer = 2
Const C_MALY As Integer = 3
Const C_ENG  As Integer = 4
Const C_CAT  As Integer = 5
Const C_CTH  As Integer = 6
Const C_EEX  As Integer = 7
Const C_INT  As Integer = 8
Const C_EF   As Integer = 9
Const C_NXT  As Integer = 10
Const C_REP  As Integer = 11
Const C_COR  As Integer = 12
Const C_WRG  As Integer = 13

' Session state
Dim mQueue()    As Long
Dim mQSize      As Long
Dim mQPos       As Long
Dim mCurRow     As Long
Dim mSessCor    As Long
Dim mSessWrg    As Long
Dim mAnswerVis  As Boolean
Dim mMode       As String    ' "ID_EN" or "EN_ID"
Dim mCategory   As String

' =============================================================================
'  BUTTON SETUP  — run once to wire up clickable shapes
' =============================================================================
Sub SetupButtons()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(CARD_SH)

    ' Remove old shapes
    Dim shp As Shape
    For Each shp In ws.Shapes
        shp.Delete
    Next shp

    ' Helper: add a rounded rectangle button
    ' AddBtn(name, macro, row1, col1, row2, col2, fillHex, textHex, label)
    Call AddBtn(ws, "btnStart",  "BtnStart",  22, 3, 22, 4, "1ABC9C", "FFFFFF", Chr(9654) & "  START / RESTART")
    Call AddBtn(ws, "btnStats",  "BtnStats",  22, 5, 22, 6, "8E44AD", "FFFFFF", Chr(128202) & "  Stats")
    Call AddBtn(ws, "btnSetup",  "BtnSetup",  22, 7, 22, 8, "7F8C8D", "FFFFFF", Chr(9881) & "  Setup")

    Call AddBtn(ws, "btnShow",   "BtnShow",   19, 3, 19, 8, "E8732A", "FFFFFF", Chr(128065) & "  Show Answer")
    Call AddBtn(ws, "btnPron",   "BtnPronounce", 20, 3, 20, 8, "D6EAF8", "1A3C5E", Chr(128266) & "  Pronounce  (Damayanti)")

    Call AddBtn(ws, "btnAgain",  "BtnAgain",  17, 3, 17, 4, "C0392B", "FFFFFF", Chr(128552) & "  Again" & Chr(10) & "(forgot)")
    Call AddBtn(ws, "btnHard",   "BtnHard",   17, 5, 17, 6, "D35400", "FFFFFF", Chr(128533) & "  Hard" & Chr(10) & "(struggled)")
    Call AddBtn(ws, "btnGood",   "BtnGood",   18, 3, 18, 4, "27AE60", "FFFFFF", Chr(128578) & "  Good" & Chr(10) & "(knew it)")
    Call AddBtn(ws, "btnEasy",   "BtnEasy",   18, 5, 18, 6, "2980B9", "FFFFFF", Chr(128516) & "  Easy" & Chr(10) & "(instant!)")

    ' Mode toggle buttons
    Call AddBtn(ws, "btnModeID", "BtnModeID", 4, 4, 4, 4, "0A5C2E", "FFFFFF", Chr(127470) & Chr(127465) & " ID" & Chr(8594) & "EN")
    Call AddBtn(ws, "btnModeEN", "BtnModeEN", 4, 5, 4, 5, "3A2A6E", "FFFFFF", Chr(127468) & Chr(127463) & " EN" & Chr(8594) & "ID")

    ' Hide rating buttons initially
    SetRatingVisible ws, False
    ws.Shapes("btnShow").Visible = False

    MsgBox "Buttons created! Click " & Chr(9654) & " START to begin.", vbInformation
End Sub

Private Sub AddBtn(ws As Worksheet, nm As String, mac As String, _
                   r1 As Long, c1 As Long, r2 As Long, c2 As Long, _
                   fillClr As String, txtClr As String, lbl As String)
    Dim t  As Double, l  As Double, w  As Double, h  As Double
    Dim rng As Range
    Set rng = ws.Range(ws.Cells(r1, c1), ws.Cells(r2, c2))
    t = rng.Top + 3
    l = rng.Left + 3
    w = rng.Width - 6
    h = rng.Height - 6

    Dim shp As Shape
    Set shp = ws.Shapes.AddShape(msoShapeRoundedRectangle, l, t, w, h)
    shp.Name = nm
    shp.OnAction = mac

    With shp.Fill
        .Visible = msoTrue
        .ForeColor.RGB = RGB(CInt("&H" & Left(fillClr, 2)), _
                             CInt("&H" & Mid(fillClr, 3, 2)), _
                             CInt("&H" & Right(fillClr, 2)))
        .Transparency = 0
    End With
    With shp.Line
        .Visible = msoFalse
    End With
    With shp.TextFrame2
        .TextRange.Text = lbl
        .TextRange.Font.Name = "Calibri"
        .TextRange.Font.Bold = msoTrue
        .TextRange.Font.Size = 11
        .TextRange.Font.Fill.ForeColor.RGB = _
            RGB(CInt("&H" & Left(txtClr, 2)), _
                CInt("&H" & Mid(txtClr, 3, 2)), _
                CInt("&H" & Right(txtClr, 2)))
        .VerticalAnchor = msoAnchorMiddle
        .TextRange.ParagraphFormat.Alignment = msoAlignCenter
        .WordWrap = msoTrue
    End With
    With shp.Shadow
        .Visible = msoTrue
        .OffsetX = 1
        .OffsetY = 2
        .Transparency = 0.7
        .Size = 100
        .Blur = 4
    End With
End Sub

' =============================================================================
'  SM-2 ALGORITHM
' =============================================================================
Sub ApplySM2(srsRow As Long, rating As Integer)
    Dim wsSRS As Worksheet
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)

    Dim interval As Double
    Dim ef       As Double
    Dim reps     As Long
    Dim grade    As Integer

    interval = Val(wsSRS.Cells(srsRow, C_INT).Value)
    ef       = Val(wsSRS.Cells(srsRow, C_EF).Value)
    reps     = CLng(wsSRS.Cells(srsRow, C_REP).Value)

    If ef < 1.3 Then ef = 2.5

    ' Map 1-4 rating → SM-2 grade 0-5
    Select Case rating
        Case 1: grade = 1   ' Again
        Case 2: grade = 3   ' Hard
        Case 3: grade = 4   ' Good
        Case 4: grade = 5   ' Easy
    End Select

    If grade < 3 Then
        reps     = 0
        interval = 1
    Else
        Select Case reps
            Case 0:    interval = 1
            Case 1:    interval = 6
            Case Else: interval = WorksheetFunction.Round(interval * ef, 0)
        End Select
        reps = reps + 1
    End If

    ' Ease factor update
    ef = ef + (0.1 - (5 - grade) * (0.08 + (5 - grade) * 0.02))
    If ef < 1.3 Then ef = 1.3
    If ef > 5   Then ef = 5

    wsSRS.Cells(srsRow, C_INT).Value = interval
    wsSRS.Cells(srsRow, C_EF).Value  = WorksheetFunction.Round(ef, 2)
    wsSRS.Cells(srsRow, C_NXT).Value = Now() + interval
    wsSRS.Cells(srsRow, C_REP).Value = reps

    If grade >= 3 Then
        wsSRS.Cells(srsRow, C_COR).Value = wsSRS.Cells(srsRow, C_COR).Value + 1
        mSessCor = mSessCor + 1
    Else
        wsSRS.Cells(srsRow, C_WRG).Value = wsSRS.Cells(srsRow, C_WRG).Value + 1
        mSessWrg = mSessWrg + 1
    End If
End Sub

' =============================================================================
'  BUILD QUEUE OF DUE CARDS (shuffled)
' =============================================================================
Sub BuildQueue()
    Dim wsSRS As Worksheet
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)

    Dim lastRow As Long
    lastRow = wsSRS.Cells(wsSRS.Rows.Count, C_NUM).End(xlUp).Row

    ' Count due cards
    Dim cnt As Long
    cnt = 0
    Dim i As Long
    For i = 2 To lastRow
        If wsSRS.Cells(i, C_NUM).Value = "" Then GoTo Skip1
        If mCategory <> "All" And wsSRS.Cells(i, C_CAT).Value <> mCategory Then GoTo Skip1
        Dim nv As Variant
        nv = wsSRS.Cells(i, C_NXT).Value
        If nv = "" Or CDate(nv) <= Now() Then cnt = cnt + 1
Skip1:
    Next i

    If cnt = 0 Then
        MsgBox "No cards due!  All " & (lastRow - 1) & " cards reviewed." & _
               Chr(10) & "Come back later, or change the category.", _
               vbInformation, Chr(127881) & " All caught up!"
        mQSize = 0
        Exit Sub
    End If

    ' Fill queue
    ReDim mQueue(1 To cnt)
    Dim q As Long
    q = 1
    For i = 2 To lastRow
        If wsSRS.Cells(i, C_NUM).Value = "" Then GoTo Skip2
        If mCategory <> "All" And wsSRS.Cells(i, C_CAT).Value <> mCategory Then GoTo Skip2
        nv = wsSRS.Cells(i, C_NXT).Value
        If nv = "" Or CDate(nv) <= Now() Then
            mQueue(q) = i
            q = q + 1
        End If
Skip2:
    Next i

    ' Fisher-Yates shuffle
    Randomize
    For i = cnt To 2 Step -1
        Dim j As Long
        j = Int(Rnd() * i) + 1
        Dim tmp As Long
        tmp = mQueue(i)
        mQueue(i) = mQueue(j)
        mQueue(j) = tmp
    Next i

    mQSize   = cnt
    mQPos    = 1
    mSessCor = 0
    mSessWrg = 0
End Sub

' =============================================================================
'  SHOW CURRENT CARD
' =============================================================================
Sub ShowCard()
    If mQSize = 0 Or mQPos > mQSize Then
        ShowSessionEnd
        Exit Sub
    End If

    Dim ws As Worksheet
    Dim wsSRS As Worksheet
    Set ws    = ThisWorkbook.Sheets(CARD_SH)
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)

    mCurRow    = mQueue(mQPos)
    mAnswerVis = False

    Dim indo As String: indo = wsSRS.Cells(mCurRow, C_INDO).Value
    Dim maly As String: maly = wsSRS.Cells(mCurRow, C_MALY).Value
    Dim eng  As String: eng  = wsSRS.Cells(mCurRow, C_ENG).Value
    Dim cat  As String: cat  = wsSRS.Cells(mCurRow, C_CAT).Value
    Dim intv As Double: intv = Val(wsSRS.Cells(mCurRow, C_INT).Value)

    Dim qWord As String
    Dim qLang As String
    If mMode = "EN_ID" Then
        qWord = eng
        qLang = Chr(127468) & Chr(127463) & "  English  " & Chr(8594) & "  Indonesian"
    Else
        qWord = indo
        qLang = Chr(127470) & Chr(127465) & "  Indonesian  " & Chr(8594) & "  English"
    End If

    ws.Range("D7").Value  = qLang
    ws.Range("D8").Value  = qWord
    ws.Range("D9").Value  = ""
    ws.Range("D11").Value = ""
    ws.Range("D12").Value = ""
    ws.Range("D13").Value = ""
    ws.Range("D14").Value = ""

    ' Progress
    ws.Range("D5").Value  = "Card " & mQPos & " of " & mQSize & _
                            "   " & Chr(9989) & " " & mSessCor & _
                            "   " & Chr(10060) & " " & mSessWrg

    ' Maturity badge
    If intv = 0 Then
        ws.Range("C5").Value = Chr(11088) & " New"
    ElseIf intv < 7 Then
        ws.Range("C5").Value = Chr(128260) & " Learning"
    ElseIf intv < 21 Then
        ws.Range("C5").Value = Chr(128170) & " Review"
    Else
        ws.Range("C5").Value = Chr(127807) & " Mature"
    End If

    ' Category
    ws.Range("D4").Value = cat

    SetRatingVisible ws, False
    ws.Shapes("btnShow").Visible = True
    ws.Shapes("btnPron").Visible = True
End Sub

' =============================================================================
'  SHOW ANSWER
' =============================================================================
Sub ShowAnswer()
    If mQSize = 0 Or mCurRow = 0 Then Exit Sub
    If mAnswerVis Then Exit Sub

    Dim ws As Worksheet
    Dim wsSRS As Worksheet
    Set ws    = ThisWorkbook.Sheets(CARD_SH)
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)

    Dim indo As String: indo = wsSRS.Cells(mCurRow, C_INDO).Value
    Dim maly As String: maly = wsSRS.Cells(mCurRow, C_MALY).Value
    Dim eng  As String: eng  = wsSRS.Cells(mCurRow, C_ENG).Value
    Dim cth  As String: cth  = wsSRS.Cells(mCurRow, C_CTH).Value
    Dim eex  As String: eex  = wsSRS.Cells(mCurRow, C_EEX).Value

    If mMode = "EN_ID" Then
        ws.Range("D11").Value = indo
        If maly <> "" And maly <> indo Then
            ws.Range("D12").Value = "Malay: " & maly
        End If
    Else
        ws.Range("D11").Value = eng
        ws.Range("D12").Value = ""
    End If

    ws.Range("D13").Value = Chr(128221) & " " & cth
    ws.Range("D14").Value = Chr(127468) & Chr(127463) & " " & eex

    mAnswerVis = True
    SetRatingVisible ws, True
    ws.Shapes("btnShow").Visible = False
End Sub

' =============================================================================
'  RATING
' =============================================================================
Sub RateCard(rating As Integer)
    If Not mAnswerVis Then
        MsgBox "Please show the answer first!", vbExclamation
        Exit Sub
    End If

    ApplySM2 mCurRow, rating

    ' If wrong, re-queue at end
    If rating = 1 Then
        mQSize = mQSize + 1
        ReDim Preserve mQueue(1 To mQSize)
        mQueue(mQSize) = mCurRow
    End If

    mQPos = mQPos + 1
    ShowCard
End Sub

Sub BtnAgain():     RateCard 1: End Sub
Sub BtnHard():      RateCard 2: End Sub
Sub BtnGood():      RateCard 3: End Sub
Sub BtnEasy():      RateCard 4: End Sub
Sub BtnShow():      ShowAnswer: End Sub

' =============================================================================
'  PRONOUNCE (macOS Damayanti voice)
' =============================================================================
Sub BtnPronounce()
    If mCurRow = 0 Or mQSize = 0 Then Exit Sub

    Dim wsSRS As Worksheet
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)

    Dim word As String
    If mMode = "EN_ID" Then
        word = wsSRS.Cells(mCurRow, C_ENG).Value
    Else
        word = wsSRS.Cells(mCurRow, C_INDO).Value
    End If

    ' Sanitise for shell
    word = Replace(word, "'",  "")
    word = Replace(word, """", "")
    word = Replace(word, "/",  " ")
    word = Trim(Split(word, "(")(0))

    Dim scr As String
    ' Try Damayanti (Indonesian), fall back to default
    scr = "do shell script ""say -v Damayanti '" & word & "' 2>/dev/null || say '" & word & "'"""

    On Error Resume Next
    MacScript scr
    On Error GoTo 0
End Sub

' =============================================================================
'  MODE TOGGLE
' =============================================================================
Sub BtnModeID()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(CARD_SH)
    mMode = "ID_EN"
    ws.Range("H3").Value = "ID_EN"
    ws.Range("D7").Value = Chr(127470) & Chr(127465) & "  Indonesian  " & Chr(8594) & "  English"
    MsgBox "Mode: Indonesian " & Chr(8594) & " English", vbInformation
End Sub

Sub BtnModeEN()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(CARD_SH)
    mMode = "EN_ID"
    ws.Range("H3").Value = "EN_ID"
    ws.Range("D7").Value = Chr(127468) & Chr(127463) & "  English  " & Chr(8594) & "  Indonesian"
    MsgBox "Mode: English " & Chr(8594) & " Indonesian", vbInformation
End Sub

' =============================================================================
'  START SESSION
' =============================================================================
Sub BtnStart()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(CARD_SH)

    ' Read config from hidden cells
    mMode     = ws.Range("H3").Value
    mCategory = ws.Range("H4").Value
    If mMode     = "" Then mMode     = "ID_EN"
    If mCategory = "" Then mCategory = "All"

    ' Ensure SRS_Data is populated
    Dim wsSRS As Worksheet
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)
    If wsSRS.Cells(2, C_NUM).Value = "" Then
        MsgBox "SRS data is empty. Run " & Chr(9881) & " Setup first.", vbExclamation
        Exit Sub
    End If

    BuildQueue
    If mQSize = 0 Then Exit Sub

    mCurRow = 0
    ShowCard
End Sub

' =============================================================================
'  SETUP — category picker
' =============================================================================
Sub BtnSetup()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(CARD_SH)

    Dim cats As String
    cats = "All" & Chr(10) & _
           "Pronouns & Determiners" & Chr(10) & _
           "Conjunctions & Connectors" & Chr(10) & _
           "Prepositions" & Chr(10) & _
           "Numbers & Quantities" & Chr(10) & _
           "Time & Frequency" & Chr(10) & _
           "Verbs - General" & Chr(10) & _
           "Verbs - Movement & Action" & Chr(10) & _
           "Verbs - Communication & Mind" & Chr(10) & _
           "Adjectives" & Chr(10) & _
           "Adverbs & Degree Words" & Chr(10) & _
           "Food & Drink" & Chr(10) & _
           "Family & Relationships" & Chr(10) & _
           "Body & Health" & Chr(10) & _
           "Nature & Environment" & Chr(10) & _
           "Places & Geography" & Chr(10) & _
           "Transport & Travel" & Chr(10) & _
           "Household & Objects" & Chr(10) & _
           "Technology & Media" & Chr(10) & _
           "Education & Knowledge" & Chr(10) & _
           "Business & Finance" & Chr(10) & _
           "Government, Law & Society" & Chr(10) & _
           "Work & Profession" & Chr(10) & _
           "Arts, Culture & Religion" & Chr(10) & _
           "Abstract Concepts & Values" & Chr(10) & _
           "Emotions & Personality" & Chr(10) & _
           "Expressions & Phrases" & Chr(10) & _
           "Colors" & Chr(10) & _
           "General"

    Dim inp As String
    inp = InputBox("Enter category to study (or 'All'):" & Chr(10) & Chr(10) & cats, _
                   "Choose Category", mCategory)

    If inp = "" Then Exit Sub

    mCategory = inp
    ws.Range("H4").Value = inp
    ws.Range("D4").Value = inp

    MsgBox "Category set to: " & inp & Chr(10) & "Click START to begin.", vbInformation
End Sub

' =============================================================================
'  STATS PANEL
' =============================================================================
Sub BtnStats()
    Dim wsSRS As Worksheet
    Set wsSRS = ThisWorkbook.Sheets(SRS_SH)

    Dim lastRow As Long
    lastRow = wsSRS.Cells(wsSRS.Rows.Count, C_NUM).End(xlUp).Row

    Dim total As Long, newC As Long, lrn As Long, mature As Long
    Dim dueNow As Long, totCor As Long, totWrg As Long

    Dim i As Long
    For i = 2 To lastRow
        If wsSRS.Cells(i, C_NUM).Value = "" Then GoTo SK
        total = total + 1
        totCor = totCor + CLng(wsSRS.Cells(i, C_COR).Value)
        totWrg = totWrg + CLng(wsSRS.Cells(i, C_WRG).Value)

        Dim intv2 As Double
        intv2 = Val(wsSRS.Cells(i, C_INT).Value)
        If intv2 = 0 Then
            newC = newC + 1
        ElseIf intv2 < 21 Then
            lrn = lrn + 1
        Else
            mature = mature + 1
        End If

        Dim nv2 As Variant
        nv2 = wsSRS.Cells(i, C_NXT).Value
        If nv2 = "" Or CDate(nv2) <= Now() Then dueNow = dueNow + 1
SK:
    Next i

    Dim acc As Long
    If (totCor + totWrg) > 0 Then acc = CLng(100 * totCor / (totCor + totWrg))

    MsgBox Chr(128202) & " Your Progress" & Chr(10) & Chr(10) & _
           "Total cards:          " & total & Chr(10) & _
           Chr(11088) & " New (unseen):       " & newC & Chr(10) & _
           Chr(128260) & " Learning (<21d):   " & lrn & Chr(10) & _
           Chr(127807) & " Mature (21d+):     " & mature & Chr(10) & Chr(10) & _
           Chr(128197) & " Due today:          " & dueNow & Chr(10) & Chr(10) & _
           Chr(9989) & "  Total correct:      " & totCor & Chr(10) & _
           Chr(10060) & " Total wrong:        " & totWrg & Chr(10) & _
           Chr(127919) & " Overall accuracy:  " & acc & "%" & Chr(10) & Chr(10) & _
           "Session: " & Chr(9989) & " " & mSessCor & "  " & Chr(10060) & " " & mSessWrg, _
           vbInformation, "Flashcard Stats"
End Sub

' =============================================================================
'  SESSION END
' =============================================================================
Sub ShowSessionEnd()
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets(CARD_SH)

    ws.Range("D7").Value  = "Session complete!"
    ws.Range("D8").Value  = Chr(127881) & " " & (mSessCor + mSessWrg) & " cards reviewed"
    ws.Range("D11").Value = Chr(9989) & " Correct: " & mSessCor & "   " & Chr(10060) & " Wrong: " & mSessWrg

    Dim acc As Long
    If (mSessCor + mSessWrg) > 0 Then
        acc = CLng(100 * mSessCor / (mSessCor + mSessWrg))
    End If
    ws.Range("D12").Value = "Accuracy: " & acc & "%  — Great work! Come back tomorrow. " & Chr(128075)
    ws.Range("D13").Value = ""
    ws.Range("D14").Value = ""
    ws.Range("C5").Value  = ""
    ws.Range("D5").Value  = "All done for now!"

    SetRatingVisible ws, False
    ws.Shapes("btnShow").Visible = False
    mQSize  = 0
    mCurRow = 0
End Sub

' =============================================================================
'  HELPER: toggle rating button visibility
' =============================================================================
Sub SetRatingVisible(ws As Worksheet, show As Boolean)
    Dim names() As String
    names = Split("btnAgain,btnHard,btnGood,btnEasy", ",")
    Dim n As Variant
    For Each n In names
        On Error Resume Next
        ws.Shapes(CStr(n)).Visible = show
        On Error GoTo 0
    Next n
End Sub
'''

# ── INJECT VBA WITH XLWINGS ───────────────────────────────────────────────────
print("\nInjecting VBA via xlwings…")
try:
    import xlwings as xw

    # Open the saved xlsm in Excel
    app = xw.App(visible=False)
    app.display_alerts = False
    wb_xw = app.books.open(DEST)

    # Add VBA module
    vba_project = wb_xw.api.VBProject
    # Remove existing module if present
    for comp in vba_project.VBComponents:
        if comp.Name == "modAnki":
            vba_project.VBComponents.Remove(comp)
            break

    new_module = vba_project.VBComponents.Add(1)   # 1 = vbext_ct_StdModule
    new_module.Name = "modAnki"
    new_module.CodeModule.AddFromString(VBA_CODE)

    wb_xw.save()
    wb_xw.close()
    app.quit()
    print("  ✓ VBA injected successfully!")
    print("\n  ► Open Bahasa_Vocab.xlsm → go to 'Flashcard' tab → click ⚙ Setup button once")
    print("    This creates all clickable buttons. Then click ▶ START to study!\n")

except Exception as e:
    print(f"  ⚠ xlwings injection failed: {e}")
    # Write VBA to .bas file as fallback
    bas_path = '/Users/wongshennan/Documents/personal/languages/indonesian/anki_vba.bas'
    with open(bas_path, 'w') as f:
        f.write("Attribute VB_Name = \"modAnki\"\n")
        f.write(VBA_CODE)
    print(f"  VBA saved to: {bas_path}")
    print("  Manual setup: Open Bahasa_Vocab.xlsm → Alt+F11 → Insert → Module → paste anki_vba.bas")

print("\nDone! File: Bahasa_Vocab.xlsm")
